#!/usr/bin/env python3
"""知产管家 · IP Keeper — 本地服务器"""
import os, json, sqlite3, mimetypes, math, zipfile, re, shutil, uuid, threading, calendar
from datetime import datetime, date, timedelta
from pathlib import Path
from io import BytesIO
from flask import Flask, request, jsonify, send_file, send_from_directory

BASE_DIR   = Path(__file__).parent
DB_PATH    = BASE_DIR / "ipkeeper.db"
UPLOAD_DIR  = BASE_DIR / "uploads"
STATIC_DIR  = BASE_DIR / "static"
WATCH_DIR   = BASE_DIR / "待关联"
UPLOAD_DIR.mkdir(exist_ok=True)
WATCH_DIR.mkdir(exist_ok=True)

# ── 启动诊断 ────────────────────────────────────────────────
print(f"ℹ️  BASE_DIR   = {BASE_DIR}")
print(f"ℹ️  STATIC_DIR = {STATIC_DIR}")
print(f"ℹ️  index.html 存在: {(STATIC_DIR/'index.html').exists()}")
if not (STATIC_DIR/'index.html').exists():
    print(f"❌  错误：static/index.html 不存在！")
    print(f"   请确保解压后的所有文件在同一个目录")

app = Flask(__name__, static_folder=str(STATIC_DIR))
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024

# CORS 配置 — 默认允许本地，可通过环境变量 IPKEEPER_CORS_ORIGINS 自定义（逗号分隔）
CORS_ORIGINS = os.environ.get('IPKEEPER_CORS_ORIGINS', '*')

ALLOWED_EXT = {'.pdf','.png','.jpg','.jpeg','.docx','.doc','.xlsx','.xls',
               '.zip','.rar','.txt','.ofd','.msg','.eml'}

# ── PDF 文书识别规则 ──────────────────────────────────────────────────────────
# 每条规则：(关键词列表, 文书名称, 建议专利状态, 建议商标状态)
# 关键词全部命中（AND）才触发，用优先级顺序排列
DOC_RULES = [
    # ── 专利 ──
    (['授权通知书'],         '授权通知书',   '已授权', None),
    (['专利证书', '授权'],   '专利证书',     '已授权', None),
    (['驳回决定'],           '驳回决定书',   '已驳回', None),
    (['复审请求', '专利'],   '复审请求书',   '复审中', None),
    (['复审决定', '专利'],   '复审决定书',   '已授权', None),  # 复审成功
    (['无效宣告', '专利'],   '无效宣告请求书', '失效', None),
    (['视为撤回'],           '视撤通知书',   '主动撤回', None),
    (['放弃专利权'],         '放弃通知书',   '失效', None),
    (['进入实质审查'],       '进入实审通知', '实审中', None),
    (['实质审查请求'],       '实审请求书',   '实审中', None),
    (['审查意见通知书', '专利'], '审查意见通知书', '实审中', None),
    (['专利申请公布'],       '申请公布通知', '已公开', None),
    (['公开号'],             '申请公布通知', '已公开', None),
    (['专利申请受理'],       '专利受理通知书', '已受理', None),
    (['申请号', '受理'],     '受理通知书',   '已受理', None),
    # ── 商标 ──
    (['商标注册证'],         '商标注册证',   None, '已注册'),
    (['商标注册申请初步审定公告'], '初审公告', None, '初审公告'),
    (['初步审定'],           '初审公告',     None, '初审公告'),
    (['不予注册', '商标'],   '驳回通知书',   None, '复审中'),
    (['驳回复审', '商标'],   '驳回复审申请', None, '复审中'),
    (['异议申请', '商标'],   '异议申请书',   None, '异议中'),
    (['异议决定', '商标'],   '异议裁定书',   None, '异议中'),
    (['撤销申请', '连续三年'], '撤三申请',   None, '撤三中'),
    (['无效宣告', '商标'],   '无效宣告申请', None, '无效中'),
    (['商标注册申请受理'],   '商标受理通知书', None, '已受理'),
    (['商标', '申请号', '受理'], '受理通知书', None, '已受理'),
]

def extract_pdf_text(path: Path, max_chars=8000) -> str:
    """提取 PDF 文本，失败返回空字符串"""
    try:
        from pdfminer.high_level import extract_text
        text = extract_text(str(path))
        return (text or '')[:max_chars]
    except Exception:
        pass
    try:
        import pypdf
        reader = pypdf.PdfReader(str(path))
        parts = []
        for page in reader.pages[:8]:
            parts.append(page.extract_text() or '')
        return '\n'.join(parts)[:max_chars]
    except Exception:
        return ''

def analyze_doc_text(text: str, item_type: str):
    """
    对提取的文本运行规则匹配。
    返回 {'doc_type': str, 'suggested_status': str|None, 'confidence': float, 'snippet': str}
    """
    if not text.strip():
        return None
    # 归一化
    t = text.replace('\n', ' ').replace('\r', ' ')
    for keywords, doc_name, patent_status, tm_status in DOC_RULES:
        if all(kw in t for kw in keywords):
            suggested = patent_status if item_type == 'patent' else tm_status
            # 取前200字作为摘要
            snippet = re.sub(r'\s+', ' ', t[:200]).strip()
            return {
                'doc_type': doc_name,
                'suggested_status': suggested,
                'confidence': 0.9,
                'snippet': snippet,
            }
    return {
        'doc_type': '其他文书',
        'suggested_status': None,
        'confidence': 0.0,
        'snippet': re.sub(r'\s+', ' ', t[:200]).strip(),
    }


FEE_TABLE = {
    # 2023年最新专利年费标准
    '发明':    {1:900,2:900,3:900,4:1200,5:1200,6:1200,7:2000,8:2000,9:2000,
               10:4000,11:4000,12:4000,13:6000,14:6000,15:6000,
               16:8000,17:8000,18:8000,19:8000,20:8000},
    '实用新型':{1:600,2:600,3:600,4:900,5:900,6:1200,7:1200,8:1200,9:2000,10:2000},
    '外观设计':{1:600,2:600,3:600,4:900,5:900,6:1200,7:1200,8:1200,
               9:2000,10:2000,11:3000,12:3000,13:3000,14:3000,15:3000},
}
FEE_REDUCTION = {
    '无费减':1.0,
    '有费减-单个主体':0.15,   # 85% 减缴
    '有费减-多个主体':0.30,   # 70% 减缴
    # 兼容旧数据
    '大型企业':1.0,'中型企业':0.6,'小型企业':0.3,'微型企业':0.15,'个人':0.15,  # 兼容旧数据（已废弃）
}

def get_fee(patent_type, year, entity):
    base = FEE_TABLE.get(patent_type, {}).get(year, 0)
    return base, math.ceil(base * FEE_REDUCTION.get(entity, 1.0))

def get_db():
    conn = sqlite3.connect(str(DB_PATH))
    conn.row_factory = sqlite3.Row
    return conn

# ── 自动关联辅助 ────────────────────────────────────────────
def _match_file(fname, db):
    """尝试从文件名/路径中提取编号并匹配知产记录。
    支持子目录：待关联/专利/、待关联/商标/、待关联/软著/
    也支持平铺命名，如 CN202110000001_授权证书.pdf
    返回 (item_type, item_dict) 或 (None, None)
    """
    name = Path(fname).name  # just filename

    # 1. Patent – CN + 12-18 digits (may have dots, hyphens)
    for pat in [r'(CN\d{10,16}(?:\.\d)?)', r'(ZL\d{10,16})', r'(PCT[/A-Z]{2}\d{4}/\d{6})']:
        m = re.search(pat, name, re.I)
        if m:
            q = f"%{m.group(1).replace('.','%')}%"
            row = db.execute("SELECT * FROM patents WHERE app_no LIKE ? OR grant_no LIKE ? OR pub_no LIKE ?",
                             [q,q,q]).fetchone()
            if row: return 'patent', row_to_dict(row)

    # 2. Trademark – 8-digit number (Chinese trademark app no)
    m = re.search(r'(?<!\d)(\d{7,9})(?!\d)', name)
    if m:
        row = db.execute("SELECT * FROM trademarks WHERE app_no=? OR reg_no=?",
                         [m.group(1), m.group(1)]).fetchone()
        if row: return 'trademark', row_to_dict(row)

    # 3. Copyright – YYYY SR/sr digits
    m = re.search(r'(\d{4}SR\d+)', name, re.I)
    if m:
        row = db.execute("SELECT * FROM copyrights WHERE reg_no LIKE ?",
                         [f"%{m.group(1)}%"]).fetchone()
        if row: return 'copyright', row_to_dict(row)

    # 4. Subfolder hints:  待关联/专利/…  or  待关联/商标/…
    path_lower = fname.lower()
    if '专利' in fname or 'patent' in path_lower:
        hint = 'patent'
    elif '商标' in fname or 'trademark' in path_lower:
        hint = 'trademark'
    elif '软著' in fname or '著作' in fname or 'copyright' in path_lower:
        hint = 'copyright'
    else:
        hint = None
    return hint, None          # matched type but no specific record


def _doc_type_from_name(fname):
    """从文件名猜测文件类型"""
    n = fname.lower()
    if any(k in n for k in ['授权通知','授权证', 'grant']): return '授权通知书'
    if any(k in n for k in ['专利证书','证书','certificate']): return '专利证书'
    if any(k in n for k in ['缴费','年费','fee','receipt','凭证']): return '缴费凭证'
    if any(k in n for k in ['注册证','商标证']): return '注册证书'
    if any(k in n for k in ['驳回','rejection']): return '驳回决定书'
    if any(k in n for k in ['公开','公告','publication']): return '公开公告'
    if any(k in n for k in ['复审','review']): return '复审文件'
    if any(k in n for k in ['软著','登记证','registration']): return '登记证书'
    return '其他'


def init_db():
    with get_db() as db:
        db.executescript("""
        CREATE TABLE IF NOT EXISTS patents (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            title TEXT NOT NULL, app_no TEXT, pub_no TEXT, grant_no TEXT,
            type TEXT DEFAULT '发明', country TEXT DEFAULT '中国', status TEXT DEFAULT '已受理',
            app_date TEXT, pub_date TEXT, grant_date TEXT, inventors TEXT, owner TEXT,
            agent TEXT, ipc TEXT, next_fee_date TEXT, current_year INTEGER DEFAULT 1,
            fee_entity TEXT DEFAULT '有费减-单个主体', notes TEXT, tags TEXT DEFAULT '[]',
            created_at TEXT DEFAULT (datetime('now','localtime')),
            updated_at TEXT DEFAULT (datetime('now','localtime'))
        );
        CREATE TABLE IF NOT EXISTS trademarks (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL, app_no TEXT, reg_no TEXT, type TEXT DEFAULT '文字',
            classes TEXT, goods_services TEXT, country TEXT DEFAULT '中国', status TEXT DEFAULT '已受理',
            app_date TEXT, reg_date TEXT, renewal_date TEXT, owner TEXT, agent TEXT,
            notes TEXT, tags TEXT DEFAULT '[]',
            created_at TEXT DEFAULT (datetime('now','localtime')),
            updated_at TEXT DEFAULT (datetime('now','localtime'))
        );
        CREATE TABLE IF NOT EXISTS copyrights (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL, version TEXT, reg_no TEXT, reg_date TEXT,
            completion_date TEXT, owner TEXT, type TEXT DEFAULT '原始', language TEXT,
            notes TEXT, tags TEXT DEFAULT '[]',
            created_at TEXT DEFAULT (datetime('now','localtime')),
            updated_at TEXT DEFAULT (datetime('now','localtime'))
        );
        CREATE TABLE IF NOT EXISTS documents (
            id INTEGER PRIMARY KEY AUTOINCREMENT, item_type TEXT NOT NULL,
            item_id INTEGER NOT NULL, filename TEXT NOT NULL, original_name TEXT NOT NULL,
            file_size INTEGER, mime_type TEXT, doc_type TEXT DEFAULT '其他',
            description TEXT, uploaded_at TEXT DEFAULT (datetime('now','localtime'))
        );
        CREATE TABLE IF NOT EXISTS status_history (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            item_type TEXT NOT NULL,
            item_id INTEGER NOT NULL,
            old_status TEXT,
            new_status TEXT NOT NULL,
            note TEXT DEFAULT '',
            operator TEXT DEFAULT '',
            changed_at TEXT DEFAULT (datetime('now','localtime'))
        );
        CREATE TABLE IF NOT EXISTS fee_payments (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            patent_id INTEGER NOT NULL,
            fee_year INTEGER,
            standard_amount INTEGER DEFAULT 0,
            paid_amount INTEGER DEFAULT 0,
            paid_date TEXT,
            payer TEXT DEFAULT '',
            agent TEXT DEFAULT '',
            receipt_filename TEXT DEFAULT '',
            receipt_original_name TEXT DEFAULT '',
            receipt_size INTEGER DEFAULT 0,
            note TEXT DEFAULT '',
            created_at TEXT DEFAULT (datetime('now','localtime'))
        );
        CREATE TABLE IF NOT EXISTS alerts_dismissed (
            alert_id TEXT PRIMARY KEY, dismissed_at TEXT DEFAULT (datetime('now','localtime'))
        );
        CREATE TABLE IF NOT EXISTS settings (key TEXT PRIMARY KEY, value TEXT);
        INSERT OR IGNORE INTO settings(key,value) VALUES
            ('company',''),('alert_patent_fee','90,30,7'),
            ('alert_trademark_renewal','365,180,30'),('alert_pct_phase','60,30'),
            ('default_fee_entity','有费减-单个主体');
        """)
        # WAL 模式 — 只需设置一次，提升并发读写性能
        db.execute("PRAGMA journal_mode=WAL")
        db.execute("PRAGMA foreign_keys=ON")
        # 性能索引 — 加速列表查询和去重检查
        db.execute("CREATE INDEX IF NOT EXISTS idx_patents_app_no ON patents(app_no)")
        db.execute("CREATE INDEX IF NOT EXISTS idx_patents_status ON patents(status)")
        db.execute("CREATE INDEX IF NOT EXISTS idx_patents_created ON patents(created_at)")
        db.execute("CREATE INDEX IF NOT EXISTS idx_trademarks_app_no ON trademarks(app_no)")
        db.execute("CREATE INDEX IF NOT EXISTS idx_trademarks_status ON trademarks(status)")
        db.execute("CREATE INDEX IF NOT EXISTS idx_trademarks_created ON trademarks(created_at)")
        db.execute("CREATE INDEX IF NOT EXISTS idx_copyrights_reg_no ON copyrights(reg_no)")
        db.execute("CREATE INDEX IF NOT EXISTS idx_copyrights_created ON copyrights(created_at)")
        db.execute("CREATE INDEX IF NOT EXISTS idx_documents_item ON documents(item_type, item_id)")
        db.execute("CREATE INDEX IF NOT EXISTS idx_fee_payments_patent ON fee_payments(patent_id)")
        db.execute("CREATE INDEX IF NOT EXISTS idx_fee_payments_date ON fee_payments(paid_date)")
        db.execute("CREATE INDEX IF NOT EXISTS idx_status_history_item ON status_history(item_type, item_id)")
        db.execute("CREATE INDEX IF NOT EXISTS idx_alerts_dismissed ON alerts_dismissed(alert_id)")
        for col in [("patents","current_year","INTEGER DEFAULT 1"),
                    ("patents","fee_entity","TEXT DEFAULT '有费减-单个主体'"),
                    ("patents","tags","TEXT DEFAULT '[]'"),
                    ("patents","grant_date","TEXT"),
                    ("trademarks","goods_services","TEXT DEFAULT ''"),
                    ("trademarks","logo_filename","TEXT DEFAULT ''"),
                    ("trademarks","tags","TEXT DEFAULT '[]'"),
                    ("trademarks","reg_date","TEXT"),
                    ("trademarks","renewal_date","TEXT"),
                    ("copyrights","tags","TEXT DEFAULT '[]'"),
                    ("copyrights","type","TEXT DEFAULT '原始'"),
                    ("copyrights","language","TEXT"),
                    ("documents","description","TEXT DEFAULT ''"),
                   ("trademarks","rejection_date","TEXT"),
                   ("patents","oa_date","TEXT"),
                   ("patents","rejection_date","TEXT"),
                   ("patents","grant_notice_date","TEXT"),
                   ("trademarks","pub_date","TEXT"),
                   ("trademarks","opposition_notice_date","TEXT"),
                   ("trademarks","reg_fee_notice_date","TEXT")]:
            try: db.execute(f"ALTER TABLE {col[0]} ADD COLUMN {col[1]} {col[2]}")
            except: pass
        # 修复历史数据：将旧的企业规模分类统一更新为"有费减-单个主体"（最常见场景）
        # 仅更新明显是旧分类的记录，用户可在详情页手动调整为"无费减"
        db.execute("""UPDATE patents SET fee_entity='有费减-单个主体'
                      WHERE fee_entity IN ('大型企业','中型企业','小型企业','微型企业','个人')
                         OR fee_entity IS NULL OR fee_entity=''""")
        db.execute("""UPDATE patents SET type='外观设计' WHERE type IN ('外观','外观专利','外观设计专利')""")
        db.execute("""UPDATE patents SET type='实用新型' WHERE type IN ('实用','实用新型专利')""")
        db.execute("""UPDATE patents SET status='已受理' WHERE status IN ('申请中','受理')""")
        db.execute("""UPDATE patents SET status='实审中' WHERE status IN ('审查中','实审','实质审查中')""")
        db.execute("""UPDATE patents SET status='已授权' WHERE status='授权'""")
        db.execute("""UPDATE patents SET status='已驳回' WHERE status='驳回'""")
        db.execute("""UPDATE patents SET status='主动撤回' WHERE status IN ('撤回','视撤')""")
        db.execute("""UPDATE patents SET status='失效' WHERE status IN ('无效','放弃','期满')""")
        db.execute("""UPDATE trademarks SET type='文字+图形' WHERE type IN ('组合','文字图形')""")
        db.execute("""UPDATE trademarks SET type='三维标志' WHERE type IN ('立体','三维')""")
        db.execute("""UPDATE trademarks SET status='已受理' WHERE status IN ('申请中','受理')""")
        db.execute("""UPDATE trademarks SET status='初审公告' WHERE status IN ('初审','初步审定')""")
        db.execute("""UPDATE trademarks SET status='已注册' WHERE status='注册'""")
        db.execute("""UPDATE trademarks SET status='异议中' WHERE status='异议'""")
        db.execute("""UPDATE trademarks SET status='撤三中' WHERE status='撤三'""")
        db.execute("""UPDATE trademarks SET status='无效中' WHERE status='无效'""")
        db.execute("""UPDATE trademarks SET status='已撤销' WHERE status='撤销'""")
        db.execute("""UPDATE trademarks SET status='已期满' WHERE status='期满'""")
        for row in db.execute("""
            SELECT id FROM trademarks
            WHERE type<>'图形' AND COALESCE(logo_filename,'')<>''
        """).fetchall():
            clear_trademark_logo(db, row['id'])
    print("✅ 数据库初始化完成")

def row_to_dict(row):
    d = dict(row)
    for k in ('tags',):
        if k in d and d[k]:
            try: d[k] = json.loads(d[k])
            except: d[k] = []
    return d

def patent_row_to_dict(row):
    d = row_to_dict(row)
    d['fee_rule'] = foreign_patent_rule_summary(d.get('country'))
    return d

def ok(data=None, **kw): return jsonify({"ok": True, "data": data, **kw})
def err(msg, code=400): return jsonify({"ok": False, "error": msg}), code
def now_str(): return datetime.now().strftime("%Y-%m-%d %H:%M:%S")

def clear_trademark_logo(db, tid):
    row = db.execute("SELECT logo_filename FROM trademarks WHERE id=?", (tid,)).fetchone()
    if row and row['logo_filename']:
        f = UPLOAD_DIR / row['logo_filename']
        if f.exists():
            f.unlink()
    db.execute("UPDATE trademarks SET logo_filename='',updated_at=? WHERE id=?", (now_str(), tid))

@app.after_request
def add_cors(resp):
    resp.headers['Access-Control-Allow-Origin'] = CORS_ORIGINS
    resp.headers['Access-Control-Allow-Headers'] = 'Content-Type'
    resp.headers['Access-Control-Allow-Methods'] = 'GET,POST,PUT,DELETE,OPTIONS'
    return resp

@app.route('/', defaults={'path': ''})
@app.route('/<path:path>')
def serve_static(path):
    if path and (STATIC_DIR / path).exists(): return send_from_directory(str(STATIC_DIR), path)
    return send_from_directory(str(STATIC_DIR), 'index.html')

# ── 状态历史工具 ──
def record_status_change(db, item_type, item_id, old_status, new_status, note='', operator=''):
    if old_status == new_status: return
    try:
        db.execute(
            "INSERT INTO status_history (item_type,item_id,old_status,new_status,note,operator) VALUES (?,?,?,?,?,?)",
            (item_type, item_id, old_status, new_status, note, operator))
    except Exception:
        pass  # status_history table may not exist in older DBs

# ══════════════ PATENTS ══════════════
def calc_current_year(app_date: str, given_year=None) -> int:
    """根据申请日自动计算当前年度（第几年年费）"""
    if given_year and int(given_year) > 1:
        return int(given_year)
    if not app_date:
        return given_year or 1
    try:
        from datetime import date
        app_y = int(str(app_date)[:4])
        cur_y = date.today().year
        yr = cur_y - app_y + 1
        return max(1, yr)
    except Exception:
        return given_year or 1

def _to_date(value):
    if not value:
        return None
    if isinstance(value, date):
        return value
    try:
        return date.fromisoformat(str(value)[:10])
    except Exception:
        return None

def _add_years(d: date, years: int):
    if not d:
        return None
    try:
        return d.replace(year=d.year + years)
    except ValueError:
        # 处理 2 月 29 日
        return d.replace(month=2, day=28, year=d.year + years)

def _add_months(d: date, months: int):
    if not d:
        return None
    total = d.month - 1 + months
    year = d.year + total // 12
    month = total % 12 + 1
    day = min(d.day, calendar.monthrange(year, month)[1])
    return date(year, month, day)

def _month_end(d: date):
    if not d:
        return None
    return date(d.year, d.month, calendar.monthrange(d.year, d.month)[1])

def _next_anniversary_after(app_d: date, after_d: date, start_year: int = 4, term_year: int = 20):
    """返回 filing date 的下一个周年日（严格晚于 after_d）。"""
    if not app_d or not after_d:
        return None
    for year_no in range(start_year, term_year + 1):
        ann = _add_years(app_d, year_no - 1)
        if ann and ann > after_d:
            return ann
    return None

FOREIGN_FEE_RULES = {
    # 说明：
    # - basis：计算基准日
    # - mode：到期日生成方式
    # - stages / start_year：对应年费阶段
    # - grace_months：宽限期（月）
    # - early_months：可提前缴费窗口（月）
    # - note：给前端/详情页展示的规则摘要
    '美国': {
        'aliases': ['US', 'USA', 'UNITED STATES', 'UNITED STATES OF AMERICA'],
        'mode': 'milestone',
        'basis': 'grant_date',
        'stages': [3.5, 7.5, 11.5],
        'early_months': 6,
        'grace_months': 6,
        'term_year': 20,
        'note': '授权日计；3.5/7.5/11.5年缴费，提前6个月可缴，逾期6个月宽限。',
    },
    '日本': {
        'aliases': ['JP', 'JAPAN'],
        'mode': 'annual',
        'basis': 'grant_date',
        'start_year': 4,
        'term_year': 20,
        'early_months': 0,
        'grace_months': 6,
        'note': '授权日计；前3年通常在授权登记时一次缴清，第4年起按年缴，截止日按上一年度末或通知书所示。',
    },
    '韩国': {
        'aliases': ['KR', 'KOREA', 'REPUBLIC OF KOREA', 'SOUTH KOREA'],
        'mode': 'annual',
        'basis': 'grant_date',
        'start_year': 4,
        'term_year': 20,
        'early_months': 0,
        'grace_months': 6,
        'note': '登记日计；前3年在设置登记时处理，第4年起按年缴到届满。',
    },
    '新加坡': {
        'aliases': ['SG', 'SINGAPORE'],
        'mode': 'annual',
        'basis': 'app_date',
        'start_year': 4,
        'term_year': 20,
        'early_months': 3,
        'grace_months': 6,
        'note': '申请日计；第4年起按年续缴，可提前3个月，逾期6个月可补缴。',
    },
    '南非': {
        'aliases': ['ZA', 'SOUTH AFRICA'],
        'mode': 'annual',
        'basis': 'app_date',
        'start_year': 4,
        'term_year': 20,
        'early_months': 0,
        'grace_months': 6,
        'note': '申请日计；自第4年起按年缴，首缴通常对应第3个周年节点，逾期6个月内可补缴并加罚金。',
    },
    '荷兰': {
        'aliases': ['NL', 'NETHERLANDS', 'HOLLAND'],
        'mode': 'annual_month_end',
        'basis': 'app_date',
        'start_year': 4,
        'term_year': 20,
        'early_months': 0,
        'grace_months': 6,
        'note': '申请日计；第4年起按年缴，首缴为申请月月底对应的第3个周年节点，逾期6个月加50%罚金。',
    },
    '欧盟': {
        'aliases': ['EP', 'EPO', 'EU', 'EUROPEAN PATENT'],
        'mode': 'annual_month_end',
        'basis': 'app_date',
        'start_year': 3,
        'term_year': 20,
        'early_months': 3,
        'grace_months': 6,
        'note': '申请日周年所在月的月末为到期日；第3年起缴，逾期6个月加附加费。',
    },
}

_FOREIGN_RULE_LOOKUP = {}
for _canon, _rule in FOREIGN_FEE_RULES.items():
    _FOREIGN_RULE_LOOKUP[_canon.upper()] = _canon
    for _alias in _rule.get('aliases', []):
        _FOREIGN_RULE_LOOKUP[str(_alias).strip().upper()] = _canon

def _normalize_country(country):
    c = (country or '').strip()
    if not c:
        return ''
    return _FOREIGN_RULE_LOOKUP.get(c.upper(), c)

def _json_list(value, default=None):
    if default is None:
        default = []
    if value is None or value == '':
        return default
    if isinstance(value, str):
        try:
            parsed = json.loads(value)
            return parsed if isinstance(parsed, list) else default
        except Exception:
            return default
    if isinstance(value, list):
        return value
    return default

def _patent_fee_rule(country):
    return FOREIGN_FEE_RULES.get(_normalize_country(country))

def foreign_patent_rule_summary(country):
    rule = _patent_fee_rule(country)
    if not rule:
        return None
    return {
        'country': _normalize_country(country),
        'mode': rule['mode'],
        'basis': rule['basis'],
        'start_year': rule.get('start_year'),
        'stages': rule.get('stages'),
        'early_months': rule.get('early_months', 0),
        'grace_months': rule.get('grace_months', 0),
        'term_year': rule.get('term_year', 20),
        'note': rule.get('note', ''),
    }

def calc_patent_current_year(country, app_date=None, grant_date=None, given_year=None, next_fee_date=None):
    """根据国家规则计算应显示的年度/缴费阶段。

    国内专利仍沿用原有申请日年度逻辑；国外专利按国家规则保留阶段号。
    """
    if given_year not in (None, ''):
        try:
            v = float(given_year)
            return int(v) if v.is_integer() else v
        except Exception:
            pass

    country = _normalize_country(country)
    rule = _patent_fee_rule(country)
    if not rule:
        return calc_current_year(app_date, given_year)

    if rule['mode'] == 'milestone':
        return 3.5
    if rule['mode'] in ('annual', 'annual_month_end'):
        return int(rule.get('start_year', 4))
    return calc_current_year(app_date, given_year)

def calc_patent_next_fee_date(country, app_date=None, grant_date=None, current_year=None, next_fee_date=None):
    """根据国家规则计算下一次缴费绝限日。

    涉外专利按国家规则自动计算；国内专利若已有历史截止日则保留原值。
    """
    country = _normalize_country(country)
    rule = _patent_fee_rule(country)
    if not rule:
        if next_fee_date:
            return next_fee_date
        return None

    app_d = _to_date(app_date)
    grant_d = _to_date(grant_date)
    stage = None
    try:
        if current_year not in (None, ''):
            stage = float(current_year)
    except Exception:
        stage = None

    if rule['mode'] == 'milestone':
        if not grant_d:
            return None
        stage = stage or 3.5
        if stage <= 3.5:
            d = _add_months(grant_d, 42)
        elif stage <= 7.5:
            d = _add_months(grant_d, 90)
        else:
            d = _add_months(grant_d, 138)
        return d.isoformat() if d else None

    if country == '新加坡':
        if not app_d:
            return next_fee_date if next_fee_date else None
        first_due = _add_years(app_d, 4)  # 5th year renewal due date
        late_grant_cutoff = _add_months(app_d, 45)
        if grant_d and grant_d > late_grant_cutoff:
            first_due = _add_months(grant_d, 3)
        return first_due.isoformat() if first_due else None

    anchor = grant_d if rule.get('basis') == 'grant_date' else app_d
    if not anchor:
        return None

    if stage is None:
        stage = float(rule.get('start_year', 4))

    years_offset = max(0, int(round(stage)) - 1)
    d = _add_years(anchor, years_offset)
    if not d:
        return None
    if rule['mode'] == 'annual_month_end':
        d = _month_end(d)
    return d.isoformat()

def next_patent_fee_stage(country, current_year=None):
    """缴费后推进到下一阶段；国外特殊国家按其本国节奏递进。"""
    country = _normalize_country(country)
    rule = _patent_fee_rule(country)
    try:
        stage = float(current_year) if current_year not in (None, '') else None
    except Exception:
        stage = None

    if not rule:
        return (stage or 1) + 1
    if rule['mode'] == 'milestone':
        if stage is None or stage <= 3.5:
            return 7.5
        if stage <= 7.5:
            return 11.5
        return None
    if stage is None:
        stage = float(rule.get('start_year', 4))
    nxt = stage + 1
    if nxt > rule.get('term_year', 20):
        return None
    return nxt

def next_patent_fee_date_after_payment(country, app_date, grant_date, current_year, next_fee_date):
    """缴费后推算下一次缴费日。"""
    country = _normalize_country(country)
    rule = _patent_fee_rule(country)
    d = _to_date(next_fee_date)
    if not rule or not d:
        return None
    if country == '新加坡':
        app_d = _to_date(app_date)
        if not app_d:
            return None
        grant_d = _to_date(grant_date)
        late_grant_cutoff = _add_months(app_d, 45)
        first_due = _add_years(app_d, 4)
        if grant_d and grant_d > late_grant_cutoff:
            first_due = _add_months(grant_d, 3)
        if first_due and abs((d - first_due).days) <= 2:
            nxt = _next_anniversary_after(app_d, d, 5, 20)
            return nxt.isoformat() if nxt else None
        nxt = _next_anniversary_after(app_d, d, 4, 20)
        return nxt.isoformat() if nxt else None
    if rule['mode'] == 'milestone':
        return _add_years(d, 4).isoformat()
    d2 = _add_years(d, 1)
    if rule['mode'] == 'annual_month_end':
        d2 = _month_end(d2)
    return d2.isoformat()

def norm_patent_type(v):
    s = (v or '').strip()
    if '外观' in s: return '外观设计'
    if '实用' in s: return '实用新型'
    if '发明' in s: return '发明'
    return s or '发明'

def norm_patent_status(v):
    s = (v or '').strip()
    if s in ('申请中','受理','已受理'): return '已受理'
    if s in ('公开','公布','已公开'): return '已公开'
    if s in ('审查中','实审','实审中','实质审查中'): return '实审中'
    if s in ('授权','已授权'): return '已授权'
    if s in ('驳回','已驳回'): return '已驳回'
    if s in ('撤回','主动撤回','视撤'): return '主动撤回'
    if s in ('无效','放弃','期满','失效'): return '失效'
    return s or '已受理'

def norm_tm_status(v):
    s = (v or '').strip()
    if s in ('申请中','受理','已受理'): return '已受理'
    if s in ('初审','初步审定','初审公告'): return '初审公告'
    if s in ('注册','已注册'): return '已注册'
    if s in ('异议','异议中'): return '异议中'
    if s in ('撤三','撤三中'): return '撤三中'
    if s in ('无效','无效中'): return '无效中'
    if s in ('撤销','已撤销'): return '已撤销'
    if s in ('期满','已期满'): return '已期满'
    return s or '已受理'

def norm_tm_type(v):
    s = (v or '').strip()
    if s in ('组合','文字图形','文字+图形'): return '文字+图形'
    if s in ('立体','三维','三维标志'): return '三维标志'
    if '颜色' in s: return '颜色组合'
    return s or '文字'

@app.route('/api/patents', methods=['GET','POST','OPTIONS'])
def patents_api():
    if request.method == 'OPTIONS': return ok()
    if request.method == 'GET':
        q = request.args.get('q','').strip()
        status  = request.args.get('status','')
        country = request.args.get('country','').strip()
        page  = max(1, request.args.get('page', 1, type=int))
        limit = min(5000, max(1, request.args.get('limit', 1000, type=int)))
        with get_db() as db:
            sql = "SELECT * FROM patents WHERE 1=1"; params = []
            if q:
                sql += " AND (title LIKE ? OR app_no LIKE ? OR pub_no LIKE ? OR grant_no LIKE ? OR inventors LIKE ? OR owner LIKE ?)"
                p = f'%{q}%'; params += [p,p,p,p,p,p]
            if status:  sql += " AND status=?";  params.append(status)
            if country: sql += " AND country=?"; params.append(country)
            count_sql = sql.replace("SELECT *", "SELECT COUNT(*)", 1)
            total = db.execute(count_sql, params).fetchone()[0]
            rows = db.execute(sql+" ORDER BY created_at DESC LIMIT ? OFFSET ?", params + [limit, (page-1)*limit]).fetchall()
        return ok({'items': [patent_row_to_dict(r) for r in rows], 'total': total, 'page': page, 'limit': limit, 'pages': max(1, math.ceil(total/limit))})
    d = request.json
    app_date = d.get('app_date')
    country = _normalize_country(d.get('country', '中国'))
    current_year = calc_patent_current_year(country, app_date, d.get('grant_date'), d.get('current_year', 1), d.get('next_fee_date'))
    p_type = norm_patent_type(d.get('type','发明'))
    p_status = norm_patent_status(d.get('status','已受理'))
    next_fee_date = calc_patent_next_fee_date(country, app_date, d.get('grant_date'), current_year, d.get('next_fee_date'))
    with get_db() as db:
        cur = db.execute("""INSERT INTO patents
            (title,app_no,pub_no,grant_no,type,country,status,app_date,pub_date,grant_date,
             inventors,owner,agent,ipc,next_fee_date,current_year,fee_entity,notes,tags,
             oa_date,rejection_date,grant_notice_date)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (d.get('title',''),d.get('app_no',''),d.get('pub_no',''),d.get('grant_no',''),
                 p_type,country,p_status,
             app_date,d.get('pub_date'),d.get('grant_date'),
             d.get('inventors',''),d.get('owner',''),d.get('agent',''),d.get('ipc',''),
             next_fee_date,current_year,d.get('fee_entity','有费减-单个主体'),
             d.get('notes',''),json.dumps(d.get('tags',[])),
             d.get('oa_date'),d.get('rejection_date'),d.get('grant_notice_date')))
        new_id = cur.lastrowid
        record_status_change(db,'patent',new_id,None,p_status,'创建案件')
    return ok({"id": new_id})

@app.route('/api/patents/<int:pid>', methods=['GET','PUT','DELETE','OPTIONS'])
def patent_detail(pid):
    if request.method == 'OPTIONS': return ok()
    if request.method == 'GET':
        with get_db() as db: row = db.execute("SELECT * FROM patents WHERE id=?",(pid,)).fetchone()
        return ok(patent_row_to_dict(row)) if row else err("not found",404)
    if request.method == 'PUT':
        d = request.json
        with get_db() as db:
            old_row = db.execute("SELECT * FROM patents WHERE id=?",(pid,)).fetchone()
            old = row_to_dict(old_row) if old_row else {}
            def _keep(key, default=''):
                v = d.get(key)
                if v not in (None, ''):
                    return v
                ov = old.get(key)
                return ov if ov not in (None, '') else default
            app_date = _keep('app_date')
            grant_date = _keep('grant_date')
            country = _normalize_country(_keep('country', '中国'))
            current_year = calc_patent_current_year(country, app_date, grant_date, _keep('current_year', 1), _keep('next_fee_date'))
            p_type = norm_patent_type(_keep('type','发明'))
            p_status = norm_patent_status(_keep('status','已受理'))
            next_fee_date = calc_patent_next_fee_date(country, app_date, grant_date, current_year, _keep('next_fee_date'))
            old_status = old.get('status')
            old_tags = _json_list(old.get('tags'))
            new_tags = _json_list(d.get('tags'), old_tags) if d.get('tags') is not None else old_tags
            db.execute("""UPDATE patents SET title=?,app_no=?,pub_no=?,grant_no=?,type=?,country=?,status=?,
                app_date=?,pub_date=?,grant_date=?,inventors=?,owner=?,agent=?,ipc=?,next_fee_date=?,
                current_year=?,fee_entity=?,notes=?,tags=?,oa_date=?,rejection_date=?,grant_notice_date=?,updated_at=? WHERE id=?""",
                (_keep('title'),_keep('app_no'),_keep('pub_no'),_keep('grant_no'),
                 p_type,country,p_status,
                 app_date,_keep('pub_date'),grant_date,
                 _keep('inventors'),_keep('owner'),_keep('agent'),_keep('ipc'),
                 next_fee_date,current_year,_keep('fee_entity','有费减-单个主体'),
                 _keep('notes'),json.dumps(new_tags),
                 _keep('oa_date'),_keep('rejection_date'),_keep('grant_notice_date'),
                 now_str(),pid))
            record_status_change(db,'patent',pid,old_status,p_status,d.get('status_note',''))
        return ok()
    with get_db() as db:
        docs = db.execute("SELECT filename FROM documents WHERE item_type='patent' AND item_id=?",(pid,)).fetchall()
        for doc in docs:
            f = UPLOAD_DIR/doc['filename']
            if f.exists(): f.unlink()
        db.execute("DELETE FROM documents WHERE item_type='patent' AND item_id=?",(pid,))
        db.execute("DELETE FROM status_history WHERE item_type='patent' AND item_id=?",(pid,))
        # 删除缴费凭证文件
        pays = db.execute("SELECT receipt_filename FROM fee_payments WHERE patent_id=?",(pid,)).fetchall()
        for p in pays:
            if p['receipt_filename']:
                f = UPLOAD_DIR/p['receipt_filename']
                if f.exists(): f.unlink()
        db.execute("DELETE FROM fee_payments WHERE patent_id=?",(pid,))
        db.execute("DELETE FROM patents WHERE id=?",(pid,))
    return ok()

@app.route('/api/patents/bulk-fee-paid', methods=['POST','OPTIONS'])
def bulk_fee_paid():
    if request.method == 'OPTIONS': return ok()
    ids = request.json.get('ids', [])
    if not ids: return err("无效ID列表")
    updated = []
    with get_db() as db:
        for pid in ids:
            row = db.execute("SELECT * FROM patents WHERE id=?",(pid,)).fetchone()
            if not row: continue
            new_year = next_patent_fee_stage(row['country'], row['current_year'])
            new_date = next_patent_fee_date_after_payment(row['country'], row['app_date'], row['grant_date'], row['current_year'], row['next_fee_date'])
            db.execute("UPDATE patents SET current_year=?,next_fee_date=?,updated_at=? WHERE id=?",
                       (new_year, new_date, now_str(), pid))
            updated.append({"id": pid, "new_year": new_year, "new_date": new_date})
    return ok({"updated": updated})

@app.route('/api/patents/recalc-years', methods=['POST','OPTIONS'])
def recalc_years():
    """批量根据申请日重新计算所有专利的当前年度"""
    if request.method == 'OPTIONS': return ok()
    from datetime import date as _date
    cur_y = _date.today().year
    updated = 0
    with get_db() as db:
        rows = db.execute("SELECT id, app_date, current_year FROM patents WHERE country='中国' AND app_date IS NOT NULL AND app_date != ''").fetchall()
        for row in rows:
            try:
                app_y = int(str(row['app_date'])[:4])
                new_year = max(1, cur_y - app_y + 1)
                if new_year != (row['current_year'] or 1):
                    db.execute("UPDATE patents SET current_year=?,updated_at=? WHERE id=?",
                               (new_year, now_str(), row['id']))
                    updated += 1
            except Exception:
                continue
    return ok({"updated": updated, "message": f"已更新 {updated} 条专利年度数据"})


# ══════════════ TRADEMARKS ══════════════
@app.route('/api/trademarks', methods=['GET','POST','OPTIONS'])
def trademarks_api():
    if request.method == 'OPTIONS': return ok()
    if request.method == 'GET':
        q = request.args.get('q','').strip()
        status  = request.args.get('status','')
        classes = request.args.get('classes','').strip()
        country = request.args.get('country','').strip()
        page  = max(1, request.args.get('page', 1, type=int))
        limit = min(5000, max(1, request.args.get('limit', 1000, type=int)))
        with get_db() as db:
            sql = "SELECT * FROM trademarks WHERE 1=1"; params = []
            if q:
                sql += " AND (name LIKE ? OR app_no LIKE ? OR reg_no LIKE ? OR owner LIKE ?)"
                p = f'%{q}%'; params += [p,p,p,p]
            if status:  sql += " AND status=?";          params.append(status)
            if classes: sql += " AND classes LIKE ?";    params.append(f'%{classes}%')
            if country: sql += " AND country=?";         params.append(country)
            count_sql = sql.replace("SELECT *", "SELECT COUNT(*)", 1)
            total = db.execute(count_sql, params).fetchone()[0]
            rows = db.execute(sql+" ORDER BY created_at DESC LIMIT ? OFFSET ?", params + [limit, (page-1)*limit]).fetchall()
        return ok({'items': [row_to_dict(r) for r in rows], 'total': total, 'page': page, 'limit': limit, 'pages': max(1, math.ceil(total/limit))})
    d = request.json
    tm_type = norm_tm_type(d.get('type','文字'))
    tm_status = norm_tm_status(d.get('status','已受理'))
    with get_db() as db:
        cur = db.execute("""INSERT INTO trademarks
            (name,app_no,reg_no,type,classes,goods_services,country,status,app_date,reg_date,renewal_date,owner,agent,notes,tags,rejection_date,pub_date,opposition_notice_date,reg_fee_notice_date)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (d.get('name',''),d.get('app_no',''),d.get('reg_no',''),tm_type,
             d.get('classes',''),d.get('goods_services',''),d.get('country','中国'),tm_status,
             d.get('app_date'),d.get('reg_date'),d.get('renewal_date'),
             d.get('owner',''),d.get('agent',''),d.get('notes',''),json.dumps(d.get('tags',[])),
             d.get('rejection_date'),d.get('pub_date'),d.get('opposition_notice_date'),d.get('reg_fee_notice_date')))
        new_id = cur.lastrowid
        record_status_change(db,'trademark',new_id,None,tm_status,'创建案件')
    return ok({"id": new_id})

@app.route('/api/trademarks/<int:tid>', methods=['GET','PUT','DELETE','OPTIONS'])
def trademark_detail(tid):
    if request.method == 'OPTIONS': return ok()
    if request.method == 'GET':
        with get_db() as db: row = db.execute("SELECT * FROM trademarks WHERE id=?",(tid,)).fetchone()
        return ok(row_to_dict(row)) if row else err("not found",404)
    if request.method == 'PUT':
        d = request.json
        tm_type = norm_tm_type(d.get('type','文字'))
        tm_status = norm_tm_status(d.get('status','已受理'))
        with get_db() as db:
            old = db.execute("SELECT status, type, logo_filename FROM trademarks WHERE id=?",(tid,)).fetchone()
            old_status = old['status'] if old else None
            if tm_type == '图形':
                db.execute("""UPDATE trademarks SET name=?,app_no=?,reg_no=?,type=?,classes=?,goods_services=?,country=?,status=?,
                    app_date=?,reg_date=?,renewal_date=?,owner=?,agent=?,notes=?,tags=?,rejection_date=?,pub_date=?,opposition_notice_date=?,reg_fee_notice_date=?,updated_at=? WHERE id=?""",
                    (d.get('name',''),d.get('app_no',''),d.get('reg_no',''),tm_type,
                     d.get('classes',''),d.get('goods_services',''),d.get('country','中国'),tm_status,
                     d.get('app_date'),d.get('reg_date'),d.get('renewal_date'),
                     d.get('owner',''),d.get('agent',''),d.get('notes',''),
                     json.dumps(d.get('tags',[])),d.get('rejection_date'),
                     d.get('pub_date'),d.get('opposition_notice_date'),d.get('reg_fee_notice_date'),
                     now_str(),tid))
            else:
                db.execute("""UPDATE trademarks SET name=?,app_no=?,reg_no=?,type=?,classes=?,goods_services=?,country=?,status=?,
                    app_date=?,reg_date=?,renewal_date=?,owner=?,agent=?,notes=?,tags=?,rejection_date=?,pub_date=?,opposition_notice_date=?,reg_fee_notice_date=?,logo_filename='',updated_at=? WHERE id=?""",
                    (d.get('name',''),d.get('app_no',''),d.get('reg_no',''),tm_type,
                     d.get('classes',''),d.get('goods_services',''),d.get('country','中国'),tm_status,
                     d.get('app_date'),d.get('reg_date'),d.get('renewal_date'),
                     d.get('owner',''),d.get('agent',''),d.get('notes',''),
                     json.dumps(d.get('tags',[])),d.get('rejection_date'),
                     d.get('pub_date'),d.get('opposition_notice_date'),d.get('reg_fee_notice_date'),
                     now_str(),tid))
                if old and old['logo_filename']:
                    f = UPLOAD_DIR / old['logo_filename']
                    if f.exists():
                        f.unlink()
            record_status_change(db,'trademark',tid,old_status,tm_status,d.get('status_note',''))
        return ok()
    with get_db() as db:
        docs = db.execute("SELECT filename FROM documents WHERE item_type='trademark' AND item_id=?",(tid,)).fetchall()
        for doc in docs:
            f = UPLOAD_DIR/doc['filename']
            if f.exists(): f.unlink()
        db.execute("DELETE FROM documents WHERE item_type='trademark' AND item_id=?",(tid,))
        db.execute("DELETE FROM status_history WHERE item_type='trademark' AND item_id=?",(tid,))
        db.execute("DELETE FROM trademarks WHERE id=?",(tid,))
    return ok()

# ══════════════ COPYRIGHTS ══════════════
@app.route('/api/copyrights', methods=['GET','POST','OPTIONS'])
def copyrights_api():
    if request.method == 'OPTIONS': return ok()
    if request.method == 'GET':
        q = request.args.get('q','').strip()
        page  = max(1, request.args.get('page', 1, type=int))
        limit = min(5000, max(1, request.args.get('limit', 1000, type=int)))
        with get_db() as db:
            sql = "SELECT * FROM copyrights WHERE 1=1"; params = []
            if q:
                sql += " AND (name LIKE ? OR reg_no LIKE ? OR owner LIKE ?)"
                p = f'%{q}%'; params += [p,p,p]
            count_sql = sql.replace("SELECT *", "SELECT COUNT(*)", 1)
            total = db.execute(count_sql, params).fetchone()[0]
            rows = db.execute(sql+" ORDER BY created_at DESC LIMIT ? OFFSET ?", params + [limit, (page-1)*limit]).fetchall()
        return ok({'items': [row_to_dict(r) for r in rows], 'total': total, 'page': page, 'limit': limit, 'pages': max(1, math.ceil(total/limit))})
    d = request.json
    with get_db() as db:
        cur = db.execute("""INSERT INTO copyrights
            (name,version,reg_no,reg_date,completion_date,owner,type,language,notes,tags)
            VALUES (?,?,?,?,?,?,?,?,?,?)""",
            (d.get('name',''),d.get('version',''),d.get('reg_no',''),d.get('reg_date'),
             d.get('completion_date'),d.get('owner',''),d.get('type','原始'),
             d.get('language',''),d.get('notes',''),json.dumps(d.get('tags',[]))))
        new_id = cur.lastrowid
    return ok({"id": new_id})

@app.route('/api/copyrights/<int:cid>', methods=['GET','PUT','DELETE','OPTIONS'])
def copyright_detail(cid):
    if request.method == 'OPTIONS': return ok()
    if request.method == 'GET':
        with get_db() as db: row = db.execute("SELECT * FROM copyrights WHERE id=?",(cid,)).fetchone()
        return ok(row_to_dict(row)) if row else err("not found",404)
    if request.method == 'PUT':
        d = request.json
        with get_db() as db:
            db.execute("""UPDATE copyrights SET name=?,version=?,reg_no=?,reg_date=?,completion_date=?,
                owner=?,type=?,language=?,notes=?,tags=?,updated_at=? WHERE id=?""",
                (d.get('name',''),d.get('version',''),d.get('reg_no',''),d.get('reg_date'),
                 d.get('completion_date'),d.get('owner',''),d.get('type'),d.get('language',''),
                 d.get('notes',''),json.dumps(d.get('tags',[])),now_str(),cid))
        return ok()
    with get_db() as db:
        docs = db.execute("SELECT filename FROM documents WHERE item_type='copyright' AND item_id=?",(cid,)).fetchall()
        for doc in docs:
            f = UPLOAD_DIR/doc['filename']
            if f.exists(): f.unlink()
        db.execute("DELETE FROM documents WHERE item_type='copyright' AND item_id=?",(cid,))
        db.execute("DELETE FROM copyrights WHERE id=?",(cid,))
    return ok()

# ══════════════ 状态历史 ══════════════
@app.route('/api/status-history/<item_type>/<int:item_id>', methods=['GET','OPTIONS'])
def get_status_history(item_type, item_id):
    if request.method == 'OPTIONS': return ok()
    with get_db() as db:
        rows = db.execute(
            "SELECT * FROM status_history WHERE item_type=? AND item_id=? ORDER BY changed_at ASC",
            (item_type, item_id)).fetchall()
    return ok([dict(r) for r in rows])

@app.route('/api/status-history', methods=['POST','OPTIONS'])
def add_status_history():
    if request.method == 'OPTIONS': return ok()
    d = request.json
    with get_db() as db:
        cur = db.execute(
            "INSERT INTO status_history (item_type,item_id,old_status,new_status,note,operator) VALUES (?,?,?,?,?,?)",
            (d.get('item_type'),d.get('item_id'),d.get('old_status'),
             d.get('new_status'),d.get('note',''),d.get('operator','')))
    return ok({"id": cur.lastrowid})

@app.route('/api/status-history/<int:hid>', methods=['DELETE','OPTIONS'])
def delete_status_history(hid):
    if request.method == 'OPTIONS': return ok()
    with get_db() as db: db.execute("DELETE FROM status_history WHERE id=?",(hid,))
    return ok()

# ══════════════ 年费缴纳历史 ══════════════
@app.route('/api/fee-payments/<int:patent_id>', methods=['GET','OPTIONS'])
def get_fee_payments(patent_id):
    if request.method == 'OPTIONS': return ok()
    with get_db() as db:
        rows = db.execute(
            "SELECT * FROM fee_payments WHERE patent_id=? ORDER BY fee_year DESC, created_at DESC",
            (patent_id,)).fetchall()
    return ok([dict(r) for r in rows])

@app.route('/api/fee-payments', methods=['POST','OPTIONS'])
def create_fee_payment():
    if request.method == 'OPTIONS': return ok()
    d = request.json
    with get_db() as db:
        cur = db.execute("""INSERT INTO fee_payments
            (patent_id,fee_year,standard_amount,paid_amount,paid_date,payer,agent,note)
            VALUES (?,?,?,?,?,?,?,?)""",
            (d.get('patent_id'),d.get('fee_year'),d.get('standard_amount',0),
             d.get('paid_amount',0),d.get('paid_date'),d.get('payer',''),
             d.get('agent',''),d.get('note','')))
    return ok({"id": cur.lastrowid})

@app.route('/api/fee-payments/<int:pay_id>/receipt', methods=['POST','OPTIONS'])
def upload_receipt(pay_id):
    if request.method == 'OPTIONS': return ok()
    if 'file' not in request.files: return err("未选择文件")
    file = request.files['file']
    ext = Path(file.filename).suffix.lower()
    if ext not in ALLOWED_EXT: return err(f"不支持的文件类型: {ext}")
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe = "".join(c for c in f"receipt_{pay_id}_{ts}_{file.filename}" if c.isalnum() or c in '._-')
    save_path = UPLOAD_DIR / safe
    file.save(str(save_path))
    size = save_path.stat().st_size
    # 删除旧凭证
    with get_db() as db:
        old = db.execute("SELECT receipt_filename FROM fee_payments WHERE id=?",(pay_id,)).fetchone()
        if old and old['receipt_filename']:
            old_f = UPLOAD_DIR / old['receipt_filename']
            if old_f.exists(): old_f.unlink()
        db.execute("UPDATE fee_payments SET receipt_filename=?,receipt_original_name=?,receipt_size=? WHERE id=?",
                   (safe, file.filename, size, pay_id))
    return ok({"filename": safe, "original_name": file.filename, "size": size})

@app.route('/api/fee-payments/<int:pay_id>/receipt/download')
def download_receipt(pay_id):
    with get_db() as db: row = db.execute("SELECT * FROM fee_payments WHERE id=?",(pay_id,)).fetchone()
    if not row or not row['receipt_filename']: return err("无凭证",404)
    f = UPLOAD_DIR / row['receipt_filename']
    if not f.exists(): return err("文件已删除",404)
    return send_file(str(f), download_name=row['receipt_original_name'], as_attachment=True)

@app.route('/api/fee-payments/<int:pay_id>/receipt/preview')
def preview_receipt(pay_id):
    with get_db() as db: row = db.execute("SELECT * FROM fee_payments WHERE id=?",(pay_id,)).fetchone()
    if not row or not row['receipt_filename']: return err("无凭证",404)
    f = UPLOAD_DIR / row['receipt_filename']
    if not f.exists(): return err("文件已删除",404)
    mime = mimetypes.guess_type(row['receipt_original_name'])[0] or 'application/octet-stream'
    return send_file(str(f), mimetype=mime)

@app.route('/api/fee-payments/<int:pay_id>', methods=['DELETE','OPTIONS'])
def delete_fee_payment(pay_id):
    if request.method == 'OPTIONS': return ok()
    with get_db() as db:
        row = db.execute("SELECT * FROM fee_payments WHERE id=?",(pay_id,)).fetchone()
        if row and row['receipt_filename']:
            f = UPLOAD_DIR / row['receipt_filename']
            if f.exists(): f.unlink()
        db.execute("DELETE FROM fee_payments WHERE id=?",(pay_id,))
    return ok()

@app.route('/api/fee-payments/batch-download', methods=['POST','OPTIONS'])
def batch_download_receipts():
    if request.method == 'OPTIONS': return ok()
    ids = request.json.get('ids', [])
    if not ids: return err("请选择要下载的记录")
    buf = BytesIO()
    count = 0
    with get_db() as db:
        with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zf:
            for pay_id in ids:
                row = db.execute("SELECT * FROM fee_payments WHERE id=?",(pay_id,)).fetchone()
                if not row or not row['receipt_filename']: continue
                f = UPLOAD_DIR / row['receipt_filename']
                if not f.exists(): continue
                # 文件名：第X年_缴费日期_原始文件名
                yr = f"第{row['fee_year']}年" if row['fee_year'] else '未知年度'
                dt = (row['paid_date'] or '').replace('-','')[:8]
                arcname = f"{yr}_{dt}_{row['receipt_original_name']}"
                zf.write(str(f), arcname)
                count += 1
    if count == 0: return err("所选记录均无凭证文件")
    buf.seek(0)
    return send_file(buf, mimetype='application/zip', as_attachment=True,
                     download_name=f'缴费凭证_{date.today()}.zip')

# ══════════════ DOCUMENTS ══════════════
@app.route('/api/documents/<item_type>/<int:item_id>', methods=['GET','OPTIONS'])
def list_docs(item_type, item_id):
    if request.method == 'OPTIONS': return ok()
    with get_db() as db:
        rows = db.execute(
            "SELECT * FROM documents WHERE item_type=? AND item_id=? ORDER BY uploaded_at DESC",
            (item_type, item_id)).fetchall()
    return ok([dict(r) for r in rows])

EVIDENCE_TYPES = ['销售合同','销售发票','商业宣传材料','广告合同','实物照片','展会参展证明',
                  '媒体报道','获奖证书','海关进出口证明','其他使用证据']

@app.route('/api/trademarks/<int:tid>/evidence', methods=['GET','OPTIONS'])
def trademark_evidence_list(tid):
    if request.method == 'OPTIONS': return ok()
    with get_db() as db:
        rows = db.execute(
            "SELECT * FROM documents WHERE item_type='trademark_evidence' AND item_id=? ORDER BY uploaded_at DESC",(tid,)).fetchall()
    return ok([dict(r) for r in rows])

@app.route('/api/trademarks/<int:tid>/logo', methods=['POST','OPTIONS'])
def upload_logo(tid):
    if request.method == 'OPTIONS': return ok()
    if 'file' not in request.files: return err("未选择文件")
    file = request.files['file']
    with get_db() as db:
        row = db.execute("SELECT type FROM trademarks WHERE id=?", (tid,)).fetchone()
    if not row:
        return err("not found", 404)
    if norm_tm_type(row['type']) != '图形':
        return err("仅图形商标可设置商标图")
    ext = Path(file.filename).suffix.lower()
    if ext not in {'.png','.jpg','.jpeg','.gif','.webp','.svg'}: return err("仅支持图片格式")
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe = f"logo_{tid}_{ts}{ext}"
    save_path = UPLOAD_DIR / safe
    file.save(str(save_path))
    with get_db() as db:
        db.execute("UPDATE trademarks SET logo_filename=?,updated_at=? WHERE id=?",(safe,now_str(),tid))
    return ok({"logo_filename": safe})

@app.route('/api/trademarks/<int:tid>/logo', methods=['DELETE','OPTIONS'])
def delete_logo(tid):
    if request.method == 'OPTIONS': return ok()
    with get_db() as db:
        clear_trademark_logo(db, tid)
    return ok()

@app.route('/api/uploads/<path:filename>')
def serve_upload(filename):
    return send_from_directory(str(UPLOAD_DIR), filename)


@app.route('/api/documents/upload', methods=['POST','OPTIONS'])
def upload_doc():
    if request.method == 'OPTIONS': return ok()
    if 'file' not in request.files: return err("未选择文件")
    file = request.files['file']
    item_type = request.form.get('item_type'); item_id = request.form.get('item_id')
    doc_type = request.form.get('doc_type','其他'); description = request.form.get('description','')
    if not file.filename or not item_type or not item_id: return err("参数缺失")
    ext = Path(file.filename).suffix.lower()
    if ext not in ALLOWED_EXT: return err(f"不支持: {ext}")
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe = "".join(c for c in f"{item_type}_{item_id}_{ts}_{file.filename}" if c.isalnum() or c in '._-')
    save_path = UPLOAD_DIR / safe; file.save(str(save_path))
    size = save_path.stat().st_size
    mime = mimetypes.guess_type(file.filename)[0] or 'application/octet-stream'
    with get_db() as db:
        cur = db.execute(
            "INSERT INTO documents (item_type,item_id,filename,original_name,file_size,mime_type,doc_type,description) VALUES (?,?,?,?,?,?,?,?)",
            (item_type,int(item_id),safe,file.filename,size,mime,doc_type,description))
    doc_id = cur.lastrowid
    # 自动分析 PDF 文书
    analysis = None
    if ext == '.pdf':
        text = extract_pdf_text(save_path)
        analysis = analyze_doc_text(text, item_type)
    return ok({"id":doc_id,"filename":safe,"original_name":file.filename,
               "file_size":size,"analysis":analysis})

@app.route('/api/documents/<int:doc_id>/analyze', methods=['GET','OPTIONS'])
def analyze_doc(doc_id):
    """重新分析已上传的文书"""
    if request.method == 'OPTIONS': return ok()
    with get_db() as db:
        row = db.execute("SELECT * FROM documents WHERE id=?",(doc_id,)).fetchone()
    if not row: return err("不存在",404)
    f = UPLOAD_DIR / row['filename']
    if not f.exists(): return err("文件已删除",404)
    if not row['filename'].endswith('.pdf'):
        return ok({'doc_type':'非PDF文件','suggested_status':None,'confidence':0,'snippet':''})
    text = extract_pdf_text(f)
    result = analyze_doc_text(text, row['item_type'])
    return ok(result)


@app.route('/api/documents/<int:doc_id>/download')
def download_doc(doc_id):
    with get_db() as db: row = db.execute("SELECT * FROM documents WHERE id=?",(doc_id,)).fetchone()
    if not row: return err("不存在",404)
    f = UPLOAD_DIR/row['filename']
    if not f.exists(): return err("已删除",404)
    return send_file(str(f), download_name=row['original_name'], as_attachment=True)

@app.route('/api/documents/<int:doc_id>/preview')
def preview_doc(doc_id):
    with get_db() as db: row = db.execute("SELECT * FROM documents WHERE id=?",(doc_id,)).fetchone()
    if not row: return err("不存在",404)
    f = UPLOAD_DIR/row['filename']
    if not f.exists(): return err("已删除",404)
    return send_file(str(f), mimetype=row['mime_type'] or 'application/octet-stream')

@app.route('/api/documents/<int:doc_id>', methods=['GET','PUT','DELETE','OPTIONS'])
def delete_doc(doc_id):
    if request.method == 'OPTIONS': return ok()
    if request.method == 'PUT':
        d = request.get_json(force=True, silent=True) or {}
        with get_db() as db:
            try:
                db.execute("UPDATE documents SET doc_type=?, description=? WHERE id=?",
                           [d.get('doc_type','其他'), d.get('description',''), doc_id])
            except Exception:
                db.execute("UPDATE documents SET doc_type=? WHERE id=?",
                           [d.get('doc_type','其他'), doc_id])
        return ok({'updated': 1})
    with get_db() as db:
        row = db.execute("SELECT * FROM documents WHERE id=?",(doc_id,)).fetchone()
        if not row: return err("不存在",404)
        f = UPLOAD_DIR/row['filename']
        if f.exists(): f.unlink()
        db.execute("DELETE FROM documents WHERE id=?",(doc_id,))
    return ok()

# ══════════════ 多维度 Excel 导出 ══════════════
ALL_FIELDS = {
    'patents': [('title','发明名称'),('app_no','申请号'),('pub_no','公开号'),('grant_no','授权号'),('type','专利类型'),('country','国家/地区'),('status','状态'),('app_date','申请日'),('pub_date','公开日'),('grant_date','授权日'),('inventors','发明人'),('owner','权利人'),('agent','代理机构'),('ipc','技术方向'),('next_fee_date','下次年费日'),('current_year','当前年度'),('fee_entity','费减主体'),('notes','备注')],
    'trademarks': [('name','商标名称'),('app_no','申请号'),('reg_no','注册号'),('type','商标类型'),('classes','尼斯分类'),('goods_services','服务项'),('country','国家/地区'),('status','状态'),('app_date','申请日'),('reg_date','注册日'),('renewal_date','续展截止日'),('owner','权利人'),('agent','代理机构'),('notes','备注')],
    'copyrights': [('name','软件名称'),('version','版本号'),('reg_no','登记号'),('reg_date','登记日期'),('completion_date','完成日期'),('owner','著作权人'),('type','著作权类型'),('language','开发语言'),('notes','备注')],
}

def _build_excel(title, field_defs, rows):
    try: import openpyxl; from openpyxl.styles import Font,PatternFill,Alignment,Border,Side
    except: return None
    wb=openpyxl.Workbook(); ws=wb.active; ws.title=title
    hf=Font(bold=True,color='FFFFFF',size=11); hfill=PatternFill('solid',fgColor='1E40AF')
    ha=Alignment(horizontal='center',vertical='center',wrap_text=True)
    thin=Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
    afill=PatternFill('solid',fgColor='EFF6FF')
    for ci,(k,l) in enumerate(field_defs,1):
        c=ws.cell(row=1,column=ci,value=l); c.font=hf; c.fill=hfill; c.alignment=ha; c.border=thin
    for ri,row in enumerate(rows,2):
        for ci,(key,_) in enumerate(field_defs,1):
            val=row.get(key,'')
            if isinstance(val,list): val=', '.join(str(v) for v in val)
            if key=='current_year' and val: val=f"第{val}年"
            c=ws.cell(row=ri,column=ci,value=str(val) if val is not None else '')
            c.alignment=Alignment(vertical='center'); c.border=thin
            if ri%2==0: c.fill=afill
    for col in ws.columns:
        ml=max((len(str(c.value or '')) for c in col),default=8)
        ws.column_dimensions[col[0].column_letter].width=min(ml+4,45)
    ws.row_dimensions[1].height=28; ws.freeze_panes='A2'
    buf=BytesIO(); wb.save(buf); buf.seek(0); return buf

@app.route('/api/export/excel', methods=['POST','OPTIONS'])
def export_excel():
    if request.method == 'OPTIONS': return ok()
    d = request.json
    data_type=d.get('type','patents'); columns=d.get('columns',[]); filters=d.get('filters',{})
    all_fm={k:l for k,l in ALL_FIELDS.get(data_type,[])}
    all_keys=[k for k,_ in ALL_FIELDS.get(data_type,[])]
    field_defs=[(k,all_fm[k]) for k in all_keys if k in (columns or all_keys) and k in all_fm]
    with get_db() as db:
        def flike(v): return f'%{v}%'
        if data_type=='patents':
            sql="SELECT * FROM patents WHERE 1=1"; params=[]
            q=filters.get('q','').strip()
            if q:
                sql+=" AND (title LIKE ? OR app_no LIKE ? OR pub_no LIKE ? OR grant_no LIKE ? OR inventors LIKE ? OR owner LIKE ? OR ipc LIKE ?)"
                params+=[flike(q)]*7
            if filters.get('status'):  sql+=" AND status=?"; params.append(filters['status'])
            if filters.get('type'):    sql+=" AND type=?"; params.append(filters['type'])
            if filters.get('country'): sql+=" AND country=?"; params.append(filters['country'])
            if filters.get('owner'):   sql+=" AND owner LIKE ?"; params.append(flike(filters['owner']))
            if filters.get('inventor'):sql+=" AND inventors LIKE ?"; params.append(flike(filters['inventor']))
            if filters.get('fee_entity'): sql+=" AND fee_entity=?"; params.append(filters['fee_entity'])
            if filters.get('app_date_from'): sql+=" AND app_date>=?"; params.append(filters['app_date_from'])
            if filters.get('app_date_to'):   sql+=" AND app_date<=?"; params.append(filters['app_date_to'])
            if filters.get('grant_date_from'): sql+=" AND grant_date>=?"; params.append(filters['grant_date_from'])
            if filters.get('grant_date_to'):   sql+=" AND grant_date<=?"; params.append(filters['grant_date_to'])
            if filters.get('year_from'): sql+=" AND current_year>=?"; params.append(int(filters['year_from']))
            if filters.get('year_to'):   sql+=" AND current_year<=?"; params.append(int(filters['year_to']))
            today=date.today().isoformat(); fd=filters.get('fee_due','')
            if fd=='overdue': sql+=" AND next_fee_date IS NOT NULL AND next_fee_date<?"; params.append(today)
            elif fd=='7':  sql+=" AND next_fee_date IS NOT NULL AND next_fee_date>=? AND next_fee_date<=date(?,'+'||'7 days')"; params+=[today,today]
            elif fd=='30': sql+=" AND next_fee_date IS NOT NULL AND next_fee_date>=? AND next_fee_date<=date(?,'+'||'30 days')"; params+=[today,today]
            elif fd=='90': sql+=" AND next_fee_date IS NOT NULL AND next_fee_date>=? AND next_fee_date<=date(?,'+'||'90 days')"; params+=[today,today]
            rows=[row_to_dict(r) for r in db.execute(sql+" ORDER BY next_fee_date,created_at DESC",params).fetchall()]
        elif data_type=='trademarks':
            sql="SELECT * FROM trademarks WHERE 1=1"; params=[]
            q=filters.get('q','').strip()
            if q:
                sql+=" AND (name LIKE ? OR app_no LIKE ? OR reg_no LIKE ? OR owner LIKE ? OR classes LIKE ?)"
                params+=[flike(q)]*5
            if filters.get('status'):   sql+=" AND status=?"; params.append(filters['status'])
            if filters.get('type'):     sql+=" AND type=?"; params.append(filters['type'])
            if filters.get('country'):  sql+=" AND country=?"; params.append(filters['country'])
            if filters.get('owner'):    sql+=" AND owner LIKE ?"; params.append(flike(filters['owner']))
            if filters.get('classes'):  sql+=" AND classes LIKE ?"; params.append(flike(filters['classes']))
            if filters.get('app_date_from'): sql+=" AND app_date>=?"; params.append(filters['app_date_from'])
            if filters.get('app_date_to'):   sql+=" AND app_date<=?"; params.append(filters['app_date_to'])
            if filters.get('reg_date_from'):     sql+=" AND reg_date>=?"; params.append(filters['reg_date_from'])
            if filters.get('reg_date_to'):       sql+=" AND reg_date<=?"; params.append(filters['reg_date_to'])
            if filters.get('renewal_date_from'): sql+=" AND renewal_date>=?"; params.append(filters['renewal_date_from'])
            if filters.get('renewal_date_to'):   sql+=" AND renewal_date<=?"; params.append(filters['renewal_date_to'])
            rows=[row_to_dict(r) for r in db.execute(sql+" ORDER BY created_at DESC",params).fetchall()]
        else:
            sql="SELECT * FROM copyrights WHERE 1=1"; params=[]
            q=filters.get('q','').strip()
            if q:
                sql+=" AND (name LIKE ? OR reg_no LIKE ? OR owner LIKE ?)"; params+=[flike(q)]*3
            if filters.get('type'):  sql+=" AND type=?"; params.append(filters['type'])
            if filters.get('owner'): sql+=" AND owner LIKE ?"; params.append(flike(filters['owner']))
            if filters.get('reg_date_from'): sql+=" AND reg_date>=?"; params.append(filters['reg_date_from'])
            if filters.get('reg_date_to'):   sql+=" AND reg_date<=?"; params.append(filters['reg_date_to'])
            if filters.get('completion_date_from'): sql+=" AND completion_date>=?"; params.append(filters['completion_date_from'])
            if filters.get('completion_date_to'):   sql+=" AND completion_date<=?"; params.append(filters['completion_date_to'])
            rows=[row_to_dict(r) for r in db.execute(sql+" ORDER BY created_at DESC",params).fetchall()]
    if not rows: return err("没有符合条件的记录")
    FNAMES={'patents':f'专利导出_{date.today()}.xlsx','trademarks':f'商标导出_{date.today()}.xlsx','copyrights':f'软著导出_{date.today()}.xlsx'}
    buf=_build_excel({'patents':'专利列表','trademarks':'商标列表','copyrights':'软著列表'}.get(data_type,'数据'),field_defs,rows)
    if not buf: return err("openpyxl 未安装")
    return send_file(buf,mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True,download_name=FNAMES.get(data_type,'export.xlsx'))

# ══════════════ 全局搜索 ══════════════
@app.route('/api/search', methods=['GET','OPTIONS'])
def global_search():
    if request.method == 'OPTIONS': return ok()
    q = request.args.get('q','').strip()
    if len(q) < 1: return ok({'patents':[],'trademarks':[],'copyrights':[],'total':0})
    p = f'%{q}%'
    with get_db() as db:
        patents = [row_to_dict(r) for r in db.execute(
            "SELECT id,title,app_no,pub_no,grant_no,status,type,country,owner,inventors,app_date FROM patents "
            "WHERE title LIKE ? OR app_no LIKE ? OR grant_no LIKE ? OR pub_no LIKE ? OR inventors LIKE ? OR owner LIKE ? OR ipc LIKE ? "
            "ORDER BY updated_at DESC LIMIT 10", [p]*7).fetchall()]
        trademarks = [row_to_dict(r) for r in db.execute(
            "SELECT id,name,app_no,reg_no,status,type,classes,country,owner,app_date FROM trademarks "
            "WHERE name LIKE ? OR app_no LIKE ? OR reg_no LIKE ? OR owner LIKE ? OR classes LIKE ? "
            "ORDER BY updated_at DESC LIMIT 10", [p]*5).fetchall()]
        copyrights = [row_to_dict(r) for r in db.execute(
            "SELECT id,name,reg_no,owner,type,version,reg_date FROM copyrights "
            "WHERE name LIKE ? OR reg_no LIKE ? OR owner LIKE ? "
            "ORDER BY updated_at DESC LIMIT 8", [p]*3).fetchall()]
    total = len(patents) + len(trademarks) + len(copyrights)
    return ok({'patents':patents,'trademarks':trademarks,'copyrights':copyrights,'total':total})

# ══════════════ EXCEL 批量导入 ══════════════
IMPORT_FIELD_MAP = {
    'patents': {
        # 名称
        '发明名称':['title'],'名称':['title'],'专利名称':['title'],'标题':['title'],
        # 申请号
        '申请号':['app_no'],'申请号码':['app_no'],
        # 公开号
        '公开号':['pub_no'],'公告号':['pub_no'],'公开公告号':['pub_no'],
        # 授权号
        '授权号':['grant_no'],'专利号':['grant_no'],'证书号':['grant_no'],
        # 类型
        '专利类型':['type'],'类型':['type'],'案件类型':['type'],
        # 国家
        '国家':['country'],'国家/地区':['country'],'地区':['country'],'申请国':['country'],
        # 状态
        '状态':['status'],'案件状态':['status'],'当前状态':['status'],
        # 日期
        '申请日':['app_date'],'申请日期':['app_date'],'提交日':['app_date'],
        '公开日':['pub_date'],'公开日期':['pub_date'],'公布日':['pub_date'],
        '授权日':['grant_date'],'授权日期':['grant_date'],'授权公告日':['grant_date'],
        # 人员
        '发明人':['inventors'],'发明人/设计人':['inventors'],'设计人':['inventors'],'创作人':['inventors'],
        '权利人/申请人':['owner'],'权利人':['owner'],'申请人':['owner'],'专利权人':['owner'],'持有人':['owner'],
        '代理机构':['agent'],'代理事务所':['agent'],'代理人':['agent'],
        # 技术
        'IPC分类号':['ipc'],'IPC':['ipc'],'IPC分类':['ipc'],'分类号':['ipc'],'国际分类':['ipc'],'技术方向':['ipc'],
        # 年费
        '下次年费日':['next_fee_date'],'年费截止日':['next_fee_date'],'年费日':['next_fee_date'],'缴费截止日':['next_fee_date'],
        '费减主体类型':['fee_entity'],'费减类型':['fee_entity'],'费减主体':['fee_entity'],'费减':['fee_entity'],
        # 其他
        '备注':['notes'],'说明':['notes'],'备注说明':['notes'],
        '标签':['tags'],
    },
    'trademarks': {
        # 名称
        '商标名称':['name'],'名称':['name'],'商标':['name'],'品牌名称':['name'],
        # 号码
        '申请号':['app_no'],'商标申请号':['app_no'],
        '注册号':['reg_no'],'商标注册号':['reg_no'],'证书号':['reg_no'],
        # 类型
        '商标类型':['type'],'类型':['type'],'标志类型':['type'],
        # 分类
        '尼斯分类':['classes'],'商品类别':['classes'],'类别':['classes'],'分类':['classes'],'尼斯类别':['classes'],
        # 服务项
        '商品/服务项':['goods_services'],'服务项':['goods_services'],'商品及服务项目':['goods_services'],'商品服务说明':['goods_services'],
        # 地区
        '国家':['country'],'国家/地区':['country'],'申请国':['country'],'地区':['country'],
        # 状态
        '状态':['status'],'案件状态':['status'],'商标状态':['status'],
        # 日期
        '申请日':['app_date'],'申请日期':['app_date'],
        '注册日':['reg_date'],'注册日期':['reg_date'],'注册公告日':['reg_date'],
        '续展截止日':['renewal_date'],'续展日期':['renewal_date'],'有效期至':['renewal_date'],'到期日':['renewal_date'],
        # 权利人
        '权利人/申请人':['owner'],'权利人':['owner'],'申请人':['owner'],'商标权人':['owner'],'持有人':['owner'],
        '代理机构':['agent'],'代理事务所':['agent'],'代理人':['agent'],
        # 其他
        '备注':['notes'],'说明':['notes'],
        '标签':['tags'],
    },
    'copyrights': {
        # 名称
        '软件名称':['name'],'名称':['name'],'软件全称':['name'],'作品名称':['name'],
        # 版本
        '版本号':['version'],'版本':['version'],'软件版本':['version'],
        # 号码
        '登记号':['reg_no'],'著作权登记号':['reg_no'],'软著登记号':['reg_no'],'证书号':['reg_no'],
        # 日期
        '完成日期':['completion_date'],'开发完成日':['completion_date'],'创作完成日期':['completion_date'],
        '登记日期':['reg_date'],'登记日':['reg_date'],'发证日期':['reg_date'],
        # 权利人
        '著作权人':['owner'],'权利人':['owner'],'申请人':['owner'],'所有人':['owner'],
        # 类型
        '著作权类型':['type'],'类型':['type'],'软件类型':['type'],
        # 技术
        '开发语言':['language'],'语言':['language'],'编程语言':['language'],
        '开发工具':['notes'],'源代码行数':['notes'],
        # 其他
        '备注':['notes'],'说明':['notes'],
    }
}

@app.route('/api/import/preview', methods=['POST','OPTIONS'])
def import_preview():
    """解析上传的Excel，返回字段映射和预览行"""
    if request.method == 'OPTIONS': return ok()
    if 'file' not in request.files: return err("未选择文件")
    file = request.files['file']
    dtype = request.form.get('type','patents')
    ext = Path(file.filename).suffix.lower()
    if ext not in ('.xlsx','.xls','.csv'): return err("仅支持 xlsx/xls/csv 格式")

    try:
        import openpyxl, io
        def normalize_val(v):
            """规范化单元格值：日期转 YYYY-MM-DD，其余转字符串"""
            if v is None: return ''
            import datetime as _dt
            if isinstance(v, (_dt.datetime, _dt.date)):
                return v.strftime('%Y-%m-%d')
            s = str(v).strip()
            # 处理 "2021-06-15 00:00:00" 这类带时间戳的日期字符串
            if len(s) > 10 and s[10] == ' ' and s[4] == '-' and s[7] == '-':
                return s[:10]
            return s

        if ext == '.csv':
            import csv
            content = file.read().decode('utf-8-sig')
            reader = csv.DictReader(io.StringIO(content))
            raw_rows = [dict(r) for r in reader]
            headers = list(raw_rows[0].keys()) if raw_rows else []
        else:
            wb = openpyxl.load_workbook(io.BytesIO(file.read()), data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            if not rows: return err("Excel 文件为空")
            headers = [str(h).strip() if h is not None else '' for h in rows[0]]
            raw_rows = []
            for row in rows[1:]:
                if any(v is not None and str(v).strip() for v in row):
                    raw_rows.append({headers[i]: normalize_val(v) for i,v in enumerate(row) if i < len(headers)})
    except Exception as e:
        return err(f"解析失败: {str(e)}")

    fmap = IMPORT_FIELD_MAP.get(dtype, {})
    mapping = {}
    for h in headers:
        if not h: continue
        hc = h.strip()
        if hc in fmap: mapping[hc] = fmap[hc][0]

    preview = raw_rows[:5]
    return ok({'headers': headers, 'mapping': mapping, 'preview': preview, 'total': len(raw_rows), 'raw_rows': raw_rows})

@app.route('/api/import/confirm', methods=['POST','OPTIONS'])
def import_confirm():
    """根据字段映射将数据写入数据库"""
    if request.method == 'OPTIONS': return ok()
    d = request.json
    dtype  = d.get('type','patents')
    rows   = d.get('rows', [])
    mapping = d.get('mapping', {})  # {excel_header: db_field}
    if not rows: return err("无数据")

    def map_row(raw):
        r = {}
        for h, dbf in mapping.items():
            v = raw.get(h,'')
            if v: r[dbf] = v
        return r

    inserted = 0
    skipped  = 0
    duplicates = 0
    with get_db() as db:
        for raw in rows:
            r = map_row(raw)
            try:
                if dtype == 'patents':
                    # 重复检测：相同申请号已存在则跳过
                    app_no = r.get('app_no','').strip()
                    if app_no and db.execute("SELECT 1 FROM patents WHERE app_no=?", (app_no,)).fetchone():
                        duplicates += 1; continue
                    app_date = r.get('app_date')
                    country = _normalize_country(r.get('country', '中国'))
                    cy = calc_patent_current_year(country, app_date, r.get('grant_date'), r.get('current_year'), r.get('next_fee_date'))
                    next_fee_date = calc_patent_next_fee_date(country, app_date, r.get('grant_date'), cy, r.get('next_fee_date'))
                    _p_all = {
                        'title': r.get('title',''), 'app_no': app_no, 'pub_no': r.get('pub_no',''),
                        'grant_no': r.get('grant_no',''), 'type': norm_patent_type(r.get('type','发明')),
                        'country': country, 'status': norm_patent_status(r.get('status','已受理')),
                        'app_date': app_date, 'pub_date': r.get('pub_date'),
                        'grant_date': r.get('grant_date'), 'inventors': r.get('inventors',''),
                        'owner': r.get('owner',''), 'agent': r.get('agent',''), 'ipc': r.get('ipc',''),
                        'next_fee_date': next_fee_date, 'current_year': cy,
                        'fee_entity': r.get('fee_entity','有费减-单个主体'),
                        'notes': r.get('notes',''), 'tags': json.dumps([])
                    }
                    _p_cols = {row[1] for row in db.execute('PRAGMA table_info(patents)').fetchall()}
                    _p_ins = {k: v for k, v in _p_all.items() if k in _p_cols}
                    cur = db.execute(
                        'INSERT INTO patents ({}) VALUES ({})'.format(
                            ','.join(_p_ins.keys()), ','.join('?'*len(_p_ins))),
                        list(_p_ins.values()))
                    record_status_change(db,'patent',cur.lastrowid,None,_p_all['status'],'Excel导入')
                elif dtype == 'trademarks':
                    # 全量导入，不做去重
                    # 动态构建INSERT，兼容不同版本的数据库schema
                    _tm_all = {
                        'name': r.get('name',''), 'app_no': r.get('app_no',''),
                        'reg_no': r.get('reg_no',''), 'type': norm_tm_type(r.get('type','文字')),
                        'classes': r.get('classes',''), 'goods_services': r.get('goods_services',''),
                        'country': r.get('country','中国'), 'status': norm_tm_status(r.get('status','已受理')),
                        'app_date': r.get('app_date'), 'reg_date': r.get('reg_date'),
                        'renewal_date': r.get('renewal_date'), 'owner': r.get('owner',''),
                        'agent': r.get('agent',''), 'notes': r.get('notes',''), 'tags': json.dumps([])
                    }
                    _tm_cols = {row[1] for row in db.execute('PRAGMA table_info(trademarks)').fetchall()}
                    _tm_ins = {k: v for k, v in _tm_all.items() if k in _tm_cols}
                    cur = db.execute(
                        'INSERT INTO trademarks ({}) VALUES ({})'.format(
                            ','.join(_tm_ins.keys()), ','.join('?'*len(_tm_ins))),
                        list(_tm_ins.values()))
                    record_status_change(db,'trademark',cur.lastrowid,None,_tm_all['status'],'Excel导入')
                else:
                    # 软著：相同登记号已存在则跳过
                    reg_no = r.get('reg_no','').strip()
                    if reg_no and db.execute("SELECT 1 FROM copyrights WHERE reg_no=?", (reg_no,)).fetchone():
                        duplicates += 1; continue
                    db.execute("""INSERT INTO copyrights
                        (name,version,reg_no,reg_date,completion_date,owner,type,language,notes,tags)
                        VALUES (?,?,?,?,?,?,?,?,?,?)""",
                        (r.get('name',''),r.get('version',''),reg_no,
                         r.get('reg_date'),r.get('completion_date'),
                         r.get('owner',''),r.get('type',''),r.get('language',''),
                         r.get('notes',''),json.dumps([])))
                inserted += 1
            except Exception as e:
                import traceback as _tb
                err_msg = str(e)
                full_err = _tb.format_exc()
                print(f"[import_confirm] row error ({dtype}): {full_err}")
                skipped += 1
                if not hasattr(import_confirm, '_first_err'):
                    import_confirm._first_err = err_msg + ' | row: ' + str({k:v for k,v in list(r.items())[:3]})
    first_err = getattr(import_confirm, '_first_err', None)
    if hasattr(import_confirm, '_first_err'): del import_confirm._first_err
    return ok({'inserted': inserted, 'skipped': skipped, 'duplicates': duplicates,
               'first_error': first_err})

@app.route('/api/import/template/<dtype>')
def import_template(dtype):
    """下载导入模板（含全字段 + 说明 + 枚举值）"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
        from openpyxl.comments import Comment
    except: return err("openpyxl未安装")

    wb = openpyxl.Workbook()
    ws = wb.active

    # ── 字段定义：(列头, 说明注释, 示例值, 列宽) ────────────────
    if dtype == 'patents':
        ws.title = '专利导入'
        fields = [
            ('发明名称',       '必填。专利申请时的名称',                                      '一种基于深度学习的图像识别方法',   30),
            ('申请号',         '格式: CN202410000001.0 / PCT/CN2024/000001',                  'CN202410000001.0',                 22),
            ('公开号',         '格式: CN115000000A（申请公布号）',                            'CN115000000A',                     18),
            ('授权号',         '格式: CN115000000B（授权公告号）',                            'CN115000000B',                     18),
            ('专利类型',       '枚举: 发明 / 实用新型 / 外观设计',                           '发明',                             14),
            ('国家/地区',      '枚举: 中国 / 美国 / 欧洲 / 日本 / 韩国 / PCT / 其他',       '中国',                             14),
            ('状态',           '枚举: 已受理/已公开/实审中/已授权/复审中/已驳回/主动撤回/失效','已受理',                          14),
            ('申请日',         '格式: YYYY-MM-DD',                                           '2024-01-15',                       14),
            ('公开日',         '格式: YYYY-MM-DD，申请公布日期',                              '2024-07-15',                       14),
            ('授权日',         '格式: YYYY-MM-DD，授权公告日期',                              '',                                 14),
            ('发明人',         '多人用逗号分隔，如: 张三,李四,王五',                         '张三,李四',                        20),
            ('权利人/申请人',  '专利权人或申请人全称',                                        'XX科技有限公司',                   22),
            ('代理机构',       '代理事务所全称（可为空）',                                    '北京XX专利代理有限公司',           24),
            ('代理人',         '代理人姓名（可为空）',                                        '张代理',                           14),
            ('技术方向',       '格式: G06F 40/00，多个用逗号分隔，用于统计分析',              'G06F 40/00,G06N 3/08',             20),
            ('下次年费日',     '格式: YYYY-MM-DD，下次需缴年费的截止日',                     '2025-01-15',                       16),
            ('费减主体类型',   '枚举: 有费减-单个主体 / 有费减-多个主体 / 无费减',           '有费减-单个主体',                  18),
            ('备注',           '其他说明信息（可为空）',                                      '',                                 20),
        ]
        samples = {f[0]: f[2] for f in fields}

    elif dtype == 'trademarks':
        ws.title = '商标导入'
        fields = [
            ('商标名称',       '必填。商标的文字或图形名称',                                  '示例商标',                         20),
            ('申请号',         '格式: 48000001（8位纯数字）',                                 '48000001',                         16),
            ('注册号',         '已注册商标的注册号（未注册可为空）',                          '48000001',                         16),
            ('商标类型',       '枚举: 文字 / 图形 / 组合 / 立体 / 颜色组合 / 声音',          '文字',                             14),
            ('尼斯分类',       '第几类，如: 9 或 9,42（多类用逗号分隔）',                    '9',                                12),
            ('商品/服务项',    '具体商品或服务项目描述（可较长）',                            '计算机软件；人工智能软件',          30),
            ('国家/地区',      '枚举: 中国 / 美国 / 欧盟 / 日本 / 马德里 / 其他',           '中国',                             14),
            ('状态',           '枚举: 已受理/初审公告/已注册/复审中/异议中/撤三中/无效中',   '已受理',                           18),
            ('申请日',         '格式: YYYY-MM-DD',                                           '2024-01-15',                       14),
            ('注册日',         '格式: YYYY-MM-DD（已注册才填）',                             '',                                 14),
            ('续展截止日',     '格式: YYYY-MM-DD，商标有效期10年',                           '',                                 16),
            ('权利人/申请人',  '商标权利人全称',                                             'XX科技有限公司',                   22),
            ('代理机构',       '商标代理机构名称（可为空）',                                  'XX商标代理有限公司',               22),
            ('备注',           '其他说明信息（可为空）',                                      '',                                 20),
        ]
        samples = {f[0]: f[2] for f in fields}

    else:  # copyrights
        ws.title = '软著导入'
        fields = [
            ('软件名称',       '必填。软件全名称（含版本号）',                                '智能管理系统V1.0',                 26),
            ('版本号',         '软件版本，如: V1.0 / 2.3.1',                                 'V1.0',                             12),
            ('登记号',         '国家版权局登记号，如: 2024SR000001（未登记可为空）',         '2024SR000001',                     20),
            ('完成日期',       '格式: YYYY-MM-DD，软件开发完成日期',                         '2023-12-01',                       14),
            ('登记日期',       '格式: YYYY-MM-DD，版权局登记日期',                           '2024-03-01',                       14),
            ('著作权人',       '著作权所有人全称',                                           'XX科技有限公司',                   22),
            ('著作权类型',     '枚举: 软件著作权 / 美术作品 / 文字作品 / 其他',             '软件著作权',                       16),
            ('开发语言',       '主要编程语言，如: Python / Java / C++ / Go / Vue',           'Python',                           14),
            ('开发工具',       '主要开发工具或框架（可为空）',                               'PyCharm, Django',                  18),
            ('源代码行数',     '源程序行数（可为空）',                                        '50000',                            14),
            ('备注',           '其他说明信息（可为空）',                                      '',                                 20),
        ]
        samples = {f[0]: f[2] for f in fields}

    # ── 样式 ─────────────────────────────────────────────────────
    hf    = Font(bold=True, color='FFFFFF', size=11)
    hfill = PatternFill('solid', fgColor='1E40AF')
    cfill = PatternFill('solid', fgColor='DBEAFE')  # 说明行淡蓝
    sfill = PatternFill('solid', fgColor='DCFCE7')  # 示例行淡绿
    ha    = Alignment(horizontal='center', vertical='center', wrap_text=True)
    la    = Alignment(horizontal='left',   vertical='center', wrap_text=True)
    thin  = Border(left=Side(style='thin'),right=Side(style='thin'),
                   top=Side(style='thin'), bottom=Side(style='thin'))
    gray_font  = Font(color='374151', size=9, italic=True)
    green_font = Font(color='166534', size=10)

    # 第1行：字段名（标题）
    for ci, (h, _desc, _sample, _w) in enumerate(fields, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = hf; c.fill = hfill; c.alignment = ha; c.border = thin
        ws.column_dimensions[c.column_letter].width = _w

    # 第2行：说明（灰色斜体）
    for ci, (_h, desc, _s, _w) in enumerate(fields, 1):
        c = ws.cell(row=2, column=ci, value=desc)
        c.font = gray_font; c.fill = cfill; c.alignment = la; c.border = thin

    # 第3行：示例数据（绿色）
    for ci, (h, _d, sample, _w) in enumerate(fields, 1):
        c = ws.cell(row=3, column=ci, value=sample)
        c.font = green_font; c.fill = sfill; c.alignment = la; c.border = thin

    # 第4行起：空白数据行（可直接填写）
    for ri in range(4, 9):
        for ci in range(1, len(fields)+1):
            c = ws.cell(row=ri, column=ci, value='')
            c.border = thin
            if ri % 2 == 0:
                c.fill = PatternFill('solid', fgColor='F9FAFB')

    # 行高
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 36
    ws.row_dimensions[3].height = 18
    ws.freeze_panes = 'A4'

    # 加一个说明sheet
    ws2 = wb.create_sheet('填写说明')
    notes = {
        'patents': [
            ['字段名',     '必填', '格式/枚举值',                                    '说明'],
            ['发明名称',   '是',   '自由文本',                                        '与申请书一致'],
            ['申请号',     '否',   'CN202410000001.0',                               '国内专利格式，PCT另有格式'],
            ['专利类型',   '否',   '发明 / 实用新型 / 外观设计',                     '默认：发明'],
            ['国家/地区',  '否',   '中国 / 美国 / 欧洲 / 日本 / 韩国 / PCT / 其他', '默认：中国'],
            ['状态',       '否',   '已受理/已公开/实审中/已授权/复审中/已驳回/主动撤回/失效', '默认：已受理'],
            ['申请日',     '否',   'YYYY-MM-DD',                                     '用于自动计算当前年度'],
            ['发明人',     '否',   '张三,李四（逗号分隔）',                          '多人用英文逗号分隔'],
            ['费减主体类型','否',  '有费减-单个主体 / 有费减-多个主体 / 无费减',     '默认：有费减-单个主体'],
            ['技术方向',   '否',   'G06F 40/00',                                    '用于统计分析，格式如 G06F 40/00'],
        ],
        'trademarks': [
            ['字段名',     '必填', '格式/枚举值',                                    '说明'],
            ['商标名称',   '是',   '自由文本',                                        '商标名称或图形描述'],
            ['申请号',     '否',   '48000001（8位）',                                '中国商标申请号格式'],
            ['商标类型',   '否',   '文字/图形/组合/立体/颜色组合/声音',             '默认：文字'],
            ['尼斯分类',   '否',   '9 或 9,42',                                      '第几类，多类用英文逗号'],
            ['国家/地区',  '否',   '中国/美国/欧盟/日本/马德里/其他',               '默认：中国'],
            ['状态',       '否',   '已受理/初审公告/已注册/复审中/异议中/撤三中/无效中','默认：已受理'],
            ['申请日',     '否',   'YYYY-MM-DD',                                     ''],
            ['续展截止日', '否',   'YYYY-MM-DD',                                     '注册后10年，到期需续展'],
        ],
        'copyrights': [
            ['字段名',     '必填', '格式/枚举值',                          '说明'],
            ['软件名称',   '是',   '自由文本，通常含版本号',               '与登记证书名称一致'],
            ['版本号',     '否',   'V1.0',                                ''],
            ['登记号',     '否',   '2024SR000001',                        '国家版权局分配'],
            ['完成日期',   '否',   'YYYY-MM-DD',                          '开发完成日期'],
            ['登记日期',   '否',   'YYYY-MM-DD',                          '版权局登记日期'],
            ['著作权人',   '否',   '自由文本',                            '公司或个人全称'],
            ['著作权类型', '否',   '软件著作权/美术作品/文字作品/其他',   '默认：软件著作权'],
            ['开发语言',   '否',   'Python / Java / C++',                 ''],
        ],
    }.get(dtype, [])

    ws2.column_dimensions['A'].width = 16
    ws2.column_dimensions['B'].width = 8
    ws2.column_dimensions['C'].width = 42
    ws2.column_dimensions['D'].width = 28
    for ri, row in enumerate(notes, 1):
        for ci, val in enumerate(row, 1):
            c = ws2.cell(row=ri, column=ci, value=val)
            c.border = thin
            if ri == 1:
                c.font = Font(bold=True, color='FFFFFF')
                c.fill = PatternFill('solid', fgColor='1E40AF')
                c.alignment = ha
            else:
                c.alignment = la
                if ri % 2 == 0: c.fill = PatternFill('solid', fgColor='F9FAFB')

    buf = BytesIO(); wb.save(buf); buf.seek(0)
    fnames = {'patents':'专利导入模板.xlsx','trademarks':'商标导入模板.xlsx','copyrights':'软著导入模板.xlsx'}
    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=fnames.get(dtype,'导入模板.xlsx'))

# ══════════════ 到期预警 API ══════════════
@app.route('/api/alerts/upcoming', methods=['GET','OPTIONS'])
def alerts_upcoming():
    """返回结构化的到期预警数据"""
    if request.method == 'OPTIONS': return ok()
    days = int(request.args.get('days', 90))
    today = date.today().isoformat()
    cutoff = (date.today() + __import__('datetime').timedelta(days=days)).isoformat()
    result = {'overdue': [], 'urgent': [], 'soon': [], 'upcoming': []}

    def classify(deadline, item):
        if not deadline: return
        diff = (date.fromisoformat(deadline) - date.today()).days
        if diff < 0:    result['overdue'].append({**item, 'days': diff, 'deadline': deadline})
        elif diff <= 14: result['urgent'].append({**item, 'days': diff, 'deadline': deadline})
        elif diff <= 30: result['soon'].append({**item, 'days': diff, 'deadline': deadline})
        else:            result['upcoming'].append({**item, 'days': diff, 'deadline': deadline})

    with get_db() as db:
        # 专利年费
        for p in db.execute("SELECT * FROM patents WHERE status IN ('已授权','授权')", []).fetchall():
            r = row_to_dict(p)
            deadline = r.get('next_fee_date') or calc_patent_next_fee_date(r.get('country'), r.get('app_date'), r.get('grant_date'), r.get('current_year'), r.get('next_fee_date'))
            if deadline and deadline <= cutoff:
                classify(deadline, {'id':r['id'],'kind':'patent_fee','title':r['title'],'app_no':r['app_no'],'owner':r['owner'],'type':r['type'],'country':r['country'],'current_year':r['current_year'],'fee_entity':r['fee_entity']})
        # 商标续展
        for t in db.execute("SELECT * FROM trademarks WHERE status IN ('已注册','注册') AND renewal_date IS NOT NULL AND renewal_date<=?", (cutoff,)).fetchall():
            r = row_to_dict(t)
            classify(r['renewal_date'], {'id':r['id'],'kind':'tm_renewal','title':r['name'],'app_no':r['app_no'],'owner':r['owner'],'classes':r['classes'],'country':r['country']})
        # 商标驳回复审（15天期限）
        for t in db.execute("SELECT * FROM trademarks WHERE status='已驳回' AND rejection_date IS NOT NULL").fetchall():
            r = row_to_dict(t)
            try:
                deadline = (date.fromisoformat(r['rejection_date']) + timedelta(days=15)).isoformat()
                if deadline <= cutoff:
                    classify(deadline, {'id':r['id'],'kind':'tm_rejection_review','title':r['name'],'app_no':r['app_no'],'owner':r['owner'],'classes':r['classes'],'country':r['country']})
            except: pass
        # ─── 专利确权期限 ───
        # 答复审查意见（15天）
        for p in db.execute("SELECT * FROM patents WHERE status IN ('实审中','审查中') AND oa_date IS NOT NULL").fetchall():
            r = row_to_dict(p)
            try:
                deadline = (date.fromisoformat(r['oa_date']) + timedelta(days=15)).isoformat()
                if deadline <= cutoff:
                    classify(deadline, {'id':r['id'],'kind':'patent_oa_response','title':r['title'],'app_no':r['app_no'],'owner':r['owner'],'type':r['type'],'country':r['country']})
            except: pass
        # 驳回复审（15天）
        for p in db.execute("SELECT * FROM patents WHERE status='已驳回' AND rejection_date IS NOT NULL").fetchall():
            r = row_to_dict(p)
            try:
                deadline = (date.fromisoformat(r['rejection_date']) + timedelta(days=15)).isoformat()
                if deadline <= cutoff:
                    classify(deadline, {'id':r['id'],'kind':'patent_rejection_review','title':r['title'],'app_no':r['app_no'],'owner':r['owner'],'type':r['type'],'country':r['country']})
            except: pass
        # 授权登记缴费（2个月）
        for p in db.execute("SELECT * FROM patents WHERE status IN ('已授权','授权') AND grant_notice_date IS NOT NULL AND grant_date IS NULL").fetchall():
            r = row_to_dict(p)
            try:
                deadline = (date.fromisoformat(r['grant_notice_date']) + timedelta(days=60)).isoformat()
                if deadline <= cutoff:
                    classify(deadline, {'id':r['id'],'kind':'patent_grant_pay','title':r['title'],'app_no':r['app_no'],'owner':r['owner'],'type':r['type'],'country':r['country']})
            except: pass
        # 实质审查请求（3年）
        for p in db.execute("SELECT * FROM patents WHERE status IN ('已受理','初审中','公开') AND app_date IS NOT NULL").fetchall():
            r = row_to_dict(p)
            try:
                deadline = (date.fromisoformat(r['app_date']) + timedelta(days=3*365)).isoformat()
                if deadline <= cutoff:
                    classify(deadline, {'id':r['id'],'kind':'patent_exam_request','title':r['title'],'app_no':r['app_no'],'owner':r['owner'],'type':r['type'],'country':r['country']})
            except: pass
        # 专利有效期（20年）
        for p in db.execute("SELECT * FROM patents WHERE status IN ('已授权','授权') AND grant_date IS NOT NULL", []).fetchall():
            r = row_to_dict(p)
            try:
                gdt = date.fromisoformat(r['grant_date'][:10])
                expire = gdt.replace(year=gdt.year+20)
                if expire.isoformat() <= cutoff:
                    classify(expire.isoformat(), {'id':r['id'],'kind':'patent_expire','title':r['title'],'app_no':r['app_no'],'owner':r['owner'],'type':r['type'],'country':r['country']})
            except: pass
        # ─── 商标确权期限 ───
        # 初审公告异议期（3个月）
        for t in db.execute("SELECT * FROM trademarks WHERE status='初审公告' AND pub_date IS NOT NULL").fetchall():
            r = row_to_dict(t)
            try:
                deadline = (date.fromisoformat(r['pub_date']) + timedelta(days=90)).isoformat()
                if deadline <= cutoff:
                    classify(deadline, {'id':r['id'],'kind':'tm_opposition_period','title':r['name'],'app_no':r['app_no'],'owner':r['owner'],'classes':r['classes'],'country':r['country']})
            except: pass
        # 异议答辩（30天）
        for t in db.execute("SELECT * FROM trademarks WHERE status='异议中' AND opposition_notice_date IS NOT NULL").fetchall():
            r = row_to_dict(t)
            try:
                deadline = (date.fromisoformat(r['opposition_notice_date']) + timedelta(days=30)).isoformat()
                if deadline <= cutoff:
                    classify(deadline, {'id':r['id'],'kind':'tm_opposition_response','title':r['name'],'app_no':r['app_no'],'owner':r['owner'],'classes':r['classes'],'country':r['country']})
            except: pass

    for lst in result.values(): lst.sort(key=lambda x: x['days'])
    result['total'] = sum(len(v) for v in result.values())
    return ok(result)

# ══════════════ SETTINGS / ALERTS / JSON EXPORT ══════════════
@app.route('/api/settings', methods=['GET','PUT','OPTIONS'])
def settings_api():
    if request.method == 'OPTIONS': return ok()
    if request.method == 'GET':
        with get_db() as db: rows=db.execute("SELECT key,value FROM settings").fetchall()
        return ok({r['key']:r['value'] for r in rows})
    d=request.json
    with get_db() as db:
        for k,v in d.items(): db.execute("INSERT OR REPLACE INTO settings(key,value) VALUES(?,?)",(k,str(v)))
    return ok()

@app.route('/api/service/status', methods=['GET','OPTIONS'])
def service_status():
    if request.method == 'OPTIONS': return ok()
    return ok({
        'running': True,
        'pid': os.getpid(),
        'port': 5678
    })

@app.route('/api/service/shutdown', methods=['POST','OPTIONS'])
def service_shutdown():
    if request.method == 'OPTIONS': return ok()
    if request.remote_addr not in ('127.0.0.1', '::1'):
        return err('仅允许本地访问', 403)
    def _stop():
        os._exit(0)
    threading.Timer(0.25, _stop).start()
    return ok({'message': '服务即将停止'})

@app.route('/api/alerts/dismissed', methods=['GET','POST','DELETE','OPTIONS'])
def alerts_dismissed():
    if request.method == 'OPTIONS': return ok()
    if request.method == 'GET':
        with get_db() as db: rows=db.execute("SELECT alert_id FROM alerts_dismissed").fetchall()
        return ok([r['alert_id'] for r in rows])
    if request.method == 'POST':
        with get_db() as db: db.execute("INSERT OR IGNORE INTO alerts_dismissed(alert_id) VALUES(?)",(request.json.get('alert_id'),))
        return ok()
    with get_db() as db: db.execute("DELETE FROM alerts_dismissed")
    return ok()

@app.route('/api/restore', methods=['POST','OPTIONS'])
def restore_json():
    """从 JSON 备份恢复数据（追加模式，跳过 app_no/reg_no 重复项）"""
    if request.method == 'OPTIONS': return ok()
    if 'file' not in request.files: return err("未选择文件")
    try:
        data = json.loads(request.files['file'].read().decode('utf-8'))
    except Exception as e:
        return err(f"JSON 解析失败: {e}")

    result = {'patents': 0, 'trademarks': 0, 'copyrights': 0,
              'skipped_patents': 0, 'skipped_trademarks': 0, 'skipped_copyrights': 0}
    with get_db() as db:
        for p in data.get('patents', []):
            try:
                if p.get('app_no') and db.execute("SELECT 1 FROM patents WHERE app_no=?",(p['app_no'],)).fetchone():
                    result['skipped_patents'] += 1; continue
                app_date = p.get('app_date')
                country = _normalize_country(p.get('country', '中国'))
                cy = calc_patent_current_year(country, app_date, p.get('grant_date'), p.get('current_year'), p.get('next_fee_date'))
                next_fee_date = calc_patent_next_fee_date(country, app_date, p.get('grant_date'), cy, p.get('next_fee_date'))
                cur = db.execute("""INSERT INTO patents
                    (title,app_no,pub_no,grant_no,type,country,status,app_date,pub_date,grant_date,
                     inventors,owner,agent,ipc,next_fee_date,current_year,fee_entity,notes,tags)
                    VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                    (p.get('title',''),p.get('app_no',''),p.get('pub_no',''),p.get('grant_no',''),
                     norm_patent_type(p.get('type','发明')),country,norm_patent_status(p.get('status','已受理')),
                     app_date,p.get('pub_date'),p.get('grant_date'),
                     p.get('inventors',''),p.get('owner',''),p.get('agent',''),p.get('ipc',''),
                     next_fee_date,cy,p.get('fee_entity','有费减-单个主体'),
                     p.get('notes',''),json.dumps(p.get('tags') or [])))
                record_status_change(db,'patent',cur.lastrowid,None,norm_patent_status(p.get('status','已受理')),'JSON恢复')
                result['patents'] += 1
            except Exception: result['skipped_patents'] += 1

        for t in data.get('trademarks', []):
            try:
                if t.get('app_no') and db.execute("SELECT 1 FROM trademarks WHERE app_no=?",(t['app_no'],)).fetchone():
                    result['skipped_trademarks'] += 1; continue
                cur = db.execute("""INSERT INTO trademarks
                    (name,app_no,reg_no,type,classes,goods_services,country,status,
                     app_date,reg_date,renewal_date,owner,agent,notes,tags)
                    VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                    (t.get('name',''),t.get('app_no',''),t.get('reg_no',''),
                     norm_tm_type(t.get('type','文字')),t.get('classes',''),t.get('goods_services',''),
                     t.get('country','中国'),norm_tm_status(t.get('status','已受理')),
                     t.get('app_date'),t.get('reg_date'),t.get('renewal_date'),
                     t.get('owner',''),t.get('agent',''),t.get('notes',''),json.dumps(t.get('tags') or [])))
                record_status_change(db,'trademark',cur.lastrowid,None,norm_tm_status(t.get('status','已受理')),'JSON恢复')
                result['trademarks'] += 1
            except Exception: result['skipped_trademarks'] += 1

        for c in data.get('copyrights', []):
            try:
                if c.get('reg_no') and db.execute("SELECT 1 FROM copyrights WHERE reg_no=?",(c['reg_no'],)).fetchone():
                    result['skipped_copyrights'] += 1; continue
                db.execute("""INSERT INTO copyrights
                    (name,version,reg_no,reg_date,completion_date,owner,type,language,notes,tags)
                    VALUES (?,?,?,?,?,?,?,?,?,?)""",
                    (c.get('name',''),c.get('version',''),c.get('reg_no',''),
                     c.get('reg_date'),c.get('completion_date'),
                     c.get('owner',''),c.get('type',''),c.get('language',''),
                     c.get('notes',''),json.dumps(c.get('tags') or [])))
                result['copyrights'] += 1
            except Exception: result['skipped_copyrights'] += 1
    return ok(result)

@app.route('/api/export')
def export_json():
    with get_db() as db:
        patents=[row_to_dict(r) for r in db.execute("SELECT * FROM patents").fetchall()]
        trademarks=[row_to_dict(r) for r in db.execute("SELECT * FROM trademarks").fetchall()]
        copyrights=[row_to_dict(r) for r in db.execute("SELECT * FROM copyrights").fetchall()]
    data=json.dumps({"patents":patents,"trademarks":trademarks,"copyrights":copyrights,"exported_at":now_str()},ensure_ascii=False,indent=2)
    return send_file(BytesIO(data.encode()),mimetype='application/json',as_attachment=True,download_name=f'ip-backup-{date.today()}.json')

@app.route('/api/export-fees', methods=['POST','OPTIONS'])
def export_fees():
    """导出选中专利的年费缴费信息Excel（网上缴费模板格式）"""
    if request.method == 'OPTIONS': return ok()
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        return err('openpyxl未安装')
    try:
        data = request.get_json(force=True, silent=True) or {}
        ids = [int(x) for x in data.get('ids', [])]
        if not ids:
            return err('请先选择要导出的专利', 400)
        with get_db() as db:
            ph = ','.join('?'*len(ids))
            rows = db.execute(
                "SELECT id,title,app_no,type,current_year,fee_entity FROM patents WHERE id IN ({})".format(ph),
                ids
            ).fetchall()
            patents = [row_to_dict(r) for r in rows]
        id_order = {v:i for i,v in enumerate(ids)}
        patents.sort(key=lambda p: id_order.get(p['id'], 999))

        OWNER   = '北京至真健康科技股份有限公司'
        CREDIT  = '91110108MA002L8738'
        HEADERS = ['序号', '申请号/专利号/国际申请号/海牙转交编号', '业务类型',
                   '票据抬头', '统一社会信用代码', '费用种类', '外币金额',
                   '费用金额（人民币）', '备注']

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '网上缴费'

        col_widths = [6, 30, 12, 26, 22, 28, 10, 16, 12]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.row_dimensions[1].height = 28

        ws.append(HEADERS)
        hdr_fill = PatternFill(fill_type='solid', fgColor='1F4E79')
        for c in range(1, len(HEADERS) + 1):
            cell = ws.cell(1, c)
            cell.font = Font(bold=True, color='FFFFFF', name='Arial', size=10)
            cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        alt_fill = PatternFill(fill_type='solid', fgColor='EBF3FB')
        for i, p in enumerate(patents, 1):
            p_type  = p.get('type') or '发明'
            yr      = int(p.get('current_year') or 1)
            entity  = p.get('fee_entity') or '有费减-单个主体'
            _, fee  = get_fee(p_type, yr, entity)
            fee_kind = f'{p_type}专利第{yr}年年费'
            ws.append([i, p.get('app_no', ''), p_type, OWNER, CREDIT, fee_kind, '', fee, ''])
            ri = i + 1
            for c in range(1, 10):
                cell = ws.cell(ri, c)
                cell.font = Font(name='Arial', size=10)
                cell.alignment = Alignment(
                    horizontal='right' if c == 8 else ('left' if c == 2 else 'center'),
                    vertical='center')
                if i % 2 == 0:
                    cell.fill = alt_fill

        ws.freeze_panes = 'A2'
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        fname = f'网上缴费_{date.today().isoformat()}.xlsx'
        return send_file(
            buf,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=fname
        )
    except Exception as e:
        import traceback
        traceback.print_exc()
        return err(f'导出失败: {str(e)}', 500)


@app.route('/api/documents/all', methods=['GET','OPTIONS'])
def docs_all():
    """获取所有附件（含所属知产名称），支持筛选"""
    if request.method == 'OPTIONS': return ok()
    item_type = request.args.get('item_type','').strip()
    doc_type  = request.args.get('doc_type','').strip()
    date_from = request.args.get('date_from','').strip()
    date_to   = request.args.get('date_to','').strip()
    q         = request.args.get('q','').strip()
    with get_db() as db:
        sql = """
            SELECT d.id, d.item_type, d.item_id, d.filename, d.original_name,
                   d.file_size, d.mime_type, d.doc_type,
                   COALESCE(d.description, '') AS description,
                   d.uploaded_at,
                   COALESCE(p.title, t.name, c.name, '') AS item_name,
                   COALESCE(t.classes, '') AS item_class
            FROM documents d
            LEFT JOIN patents    p ON d.item_type='patent'    AND d.item_id=p.id
            LEFT JOIN trademarks t ON d.item_type='trademark' AND d.item_id=t.id
            LEFT JOIN copyrights c ON d.item_type='copyright' AND d.item_id=c.id
            WHERE 1=1
        """
        params = []
        if item_type: sql += " AND d.item_type=?"; params.append(item_type)
        if doc_type:  sql += " AND d.doc_type=?";  params.append(doc_type)
        if date_from: sql += " AND d.uploaded_at>=?"; params.append(date_from)
        if date_to:   sql += " AND d.uploaded_at<=?"; params.append(date_to + ' 23:59:59')
        if q:
            p2 = f'%{q}%'
            sql += " AND (d.original_name LIKE ? OR d.filename LIKE ? OR COALESCE(p.title,t.name,c.name,'') LIKE ? OR d.doc_type LIKE ?)"
            params += [p2,p2,p2,p2]
        sql += " ORDER BY d.uploaded_at DESC"
        rows = [dict(r) for r in db.execute(sql, params).fetchall()]
    return ok(rows)

@app.route('/api/documents/count', methods=['GET','OPTIONS'])
def docs_count():
    """返回附件总数"""
    if request.method == 'OPTIONS': return ok()
    with get_db() as db:
        row = db.execute("SELECT COUNT(*) AS c FROM documents").fetchone()
    return ok({'count': row['c'] if row else 0})

@app.route('/api/documents/bulk-download', methods=['POST','OPTIONS'])
def docs_bulk_download():
    """批量下载附件（打包ZIP）"""
    if request.method == 'OPTIONS': return ok()
    try:
        ids = [int(x) for x in (request.json or {}).get('ids', [])]
        if not ids: return err('请选择要下载的附件', 400)
        with get_db() as db:
            ph = ','.join('?'*len(ids))
            docs = [dict(r) for r in db.execute(
                f"SELECT filename, original_name FROM documents WHERE id IN ({ph})", ids
            ).fetchall()]
        buf = BytesIO()
        name_count = {}
        with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zf:
            for d in docs:
                fpath = UPLOAD_DIR / d['filename']
                if not fpath.exists(): continue
                oname = d['original_name'] or d['filename']
                # Handle duplicate filenames in zip
                base, ext = os.path.splitext(oname)
                if oname in name_count:
                    name_count[oname] += 1
                    oname = f"{base}_{name_count[oname]}{ext}"
                else:
                    name_count[oname] = 0
                zf.write(fpath, oname)
        buf.seek(0)
        fname = f'附件打包_{date.today().isoformat()}.zip'
        return send_file(buf, mimetype='application/zip', as_attachment=True, download_name=fname)
    except Exception as e:
        import traceback; traceback.print_exc()
        return err(f'打包失败: {str(e)}', 500)


# ══════════ 批量上传附件 ══════════════════════════════════════

def _match_record_by_hint(db, hint):
    """通过文件名或文件夹名提取申请号/名称，匹配知产记录"""
    if not hint: return None, None, None
    import re

    # ① 专利申请号: CN20xx...   PCT/...  ZL...
    for pat in [r'(CN[\d]{9,14}[.\d]*)', r'(PCT/[A-Z]{2}[\d]{4}/[\d]+)', r'(ZL[\d]{10,14})']:
        m = re.search(pat, hint, re.IGNORECASE)
        if m:
            val = m.group(1).upper()
            row = db.execute("SELECT id,title FROM patents WHERE UPPER(app_no) LIKE ? OR UPPER(pub_no) LIKE ? OR UPPER(grant_no) LIKE ?",
                             [f'%{val}%']*3).fetchone()
            if row: return 'patent', row['id'], row['title']

    # ② 商标申请号: 7-9位纯数字
    m = re.search(r'(?<![0-9])([0-9]{7,9})(?![0-9])', hint)
    if m:
        val = m.group(1)
        row = db.execute("SELECT id,name FROM trademarks WHERE app_no=? OR reg_no=?", [val, val]).fetchone()
        if row: return 'trademark', row['id'], row['name']

    # ③ 软著登记号: 2020SR...
    m = re.search(r'([0-9]{4}SR[0-9]+)', hint, re.IGNORECASE)
    if m:
        val = m.group(1).upper()
        row = db.execute("SELECT id,name FROM copyrights WHERE UPPER(reg_no) LIKE ?", [f'%{val}%']).fetchone()
        if row: return 'copyright', row['id'], row['name']

    # ④ 模糊名称匹配（去除扩展名和常见前缀后取前10字）
    clean = re.sub(r'[.](pdf|jpg|jpeg|png|docx|doc|xlsx|xls|zip|rar|ofd)$', '', hint, flags=re.IGNORECASE)
    clean = re.sub(r'[_ \t-]', ' ', clean).strip()
    words = [w for w in clean.split() if len(w) >= 2][:3]
    for w in words:
        lk = f'%{w}%'
        row = db.execute("SELECT id,title AS name FROM patents WHERE title LIKE ? LIMIT 1", [lk]).fetchone()
        if row: return 'patent', row['id'], row['name']
        row = db.execute("SELECT id,name FROM trademarks WHERE name LIKE ? LIMIT 1", [lk]).fetchone()
        if row: return 'trademark', row['id'], row['name']
        row = db.execute("SELECT id,name FROM copyrights WHERE name LIKE ? LIMIT 1", [lk]).fetchone()
        if row: return 'copyright', row['id'], row['name']

    return None, None, None


@app.route('/api/documents/batch-parse', methods=['POST','OPTIONS'])
def docs_batch_parse():
    """
    批量上传并自动匹配：
    - 支持直接上传多个文件 (files[])
    - 支持上传 ZIP 包（含目录结构：<申请号>/<文件名>）
    临时保存在 UPLOAD_DIR/tmp/，返回匹配预览
    """
    if request.method == 'OPTIONS': return ok()
    import uuid, shutil

    TEMP_DIR = UPLOAD_DIR / 'tmp'
    TEMP_DIR.mkdir(exist_ok=True)
    batch_id = str(uuid.uuid4())[:12]
    batch_dir = TEMP_DIR / batch_id
    batch_dir.mkdir()

    results = []
    files_to_process = []  # (temp_path, original_name, hint_from_folder)

    try:
        uploaded = request.files.getlist('files[]') or request.files.getlist('files')
        if not uploaded:
            return err('请选择要上传的文件', 400)

        for uf in uploaded:
            if not uf.filename: continue
            ext = Path(uf.filename).suffix.lower()
            # 如果是 ZIP，解压后处理
            if ext == '.zip':
                zip_path = batch_dir / '__upload__.zip'
                uf.save(str(zip_path))
                try:
                    with zipfile.ZipFile(str(zip_path), 'r') as zf:
                        for zi in zf.infolist():
                            if zi.is_dir(): continue
                            zname = zi.filename
                            # 跳过 macOS 元数据
                            if '__MACOSX' in zname or zname.startswith('.') or '/.'  in zname: continue
                            parts = Path(zname).parts
                            folder_hint = parts[-2] if len(parts) >= 2 else ''
                            fname = parts[-1]
                            if Path(fname).suffix.lower() not in ALLOWED_EXT: continue
                            dest = batch_dir / f'{len(files_to_process):04d}_{fname}'
                            with zf.open(zi) as src, open(str(dest), 'wb') as dst:
                                dst.write(src.read())
                            files_to_process.append((dest, fname, folder_hint or fname))
                except Exception as ze:
                    return err(f'ZIP解压失败: {ze}', 400)
                finally:
                    zip_path.unlink(missing_ok=True)
            else:
                if ext not in ALLOWED_EXT: continue
                dest = batch_dir / f'{len(files_to_process):04d}_{uf.filename}'
                uf.save(str(dest))
                files_to_process.append((dest, uf.filename, uf.filename))

        if not files_to_process:
            shutil.rmtree(str(batch_dir), ignore_errors=True)
            return err('未找到可处理的文件', 400)

        with get_db() as db:
            for tmp_path, orig_name, hint in files_to_process:
                item_type, item_id, item_name = _match_record_by_hint(db, hint)
                results.append({
                    'tmp': tmp_path.name,            # 服务器临时文件名
                    'original_name': orig_name,
                    'size': tmp_path.stat().st_size,
                    'item_type':  item_type or '',
                    'item_id':    item_id   or '',
                    'item_name':  item_name or '',
                    'matched':    bool(item_type and item_id),
                    'doc_type':   '其他',
                })
    except Exception as e:
        import traceback; traceback.print_exc()
        return err(f'解析失败: {e}', 500)

    return ok({'batch_id': batch_id, 'files': results})


@app.route('/api/documents/batch-commit', methods=['POST','OPTIONS'])
def docs_batch_commit():
    """
    确认批量上传：将临时文件永久存储并建立关联
    Body: {batch_id, files: [{tmp, original_name, item_type, item_id, doc_type}]}
    """
    if request.method == 'OPTIONS': return ok()
    import shutil

    d = request.json or {}
    batch_id = d.get('batch_id', '')
    files = d.get('files', [])
    if not batch_id or not files: return err('参数错误', 400)

    TEMP_DIR = UPLOAD_DIR / 'tmp' / batch_id
    if not TEMP_DIR.exists(): return err('批次已过期，请重新上传', 400)

    saved = 0; skipped = 0; errors = []
    with get_db() as db:
        for f in files:
            tmp_name = f.get('tmp', '')
            item_type = f.get('item_type', '').strip()
            item_id   = f.get('item_id', '')
            orig_name = f.get('original_name', tmp_name)
            doc_type  = f.get('doc_type', '其他')

            if not tmp_name or not item_type or not item_id:
                skipped += 1; continue

            tmp_path = TEMP_DIR / tmp_name
            if not tmp_path.exists(): skipped += 1; continue

            try:
                ext = Path(orig_name).suffix.lower()
                ts = datetime.now().strftime('%Y%m%d_%H%M%S_%f')[:19]
                safe = ''.join(c for c in f'{item_type}_{item_id}_{ts}_{orig_name}' if c.isalnum() or c in '._-')
                dest = UPLOAD_DIR / safe
                shutil.move(str(tmp_path), str(dest))
                size = dest.stat().st_size
                mime = mimetypes.guess_type(orig_name)[0] or 'application/octet-stream'
                cur = db.execute(
                    "INSERT INTO documents (item_type,item_id,filename,original_name,file_size,mime_type,doc_type,description) VALUES (?,?,?,?,?,?,?,?)",
                    (item_type, int(item_id), safe, orig_name, size, mime, doc_type, f'批量上传'))
                # PDF自动分析
                if ext == '.pdf':
                    try:
                        text = extract_pdf_text(dest)
                        analysis = analyze_doc_text(text, item_type)
                        if analysis and analysis.get('doc_type') and analysis['doc_type'] != '其他':
                            db.execute("UPDATE documents SET doc_type=? WHERE id=?", (analysis['doc_type'], cur.lastrowid))
                    except Exception: pass
                saved += 1
            except Exception as e:
                errors.append(f'{orig_name}: {e}')
                skipped += 1

    # 清理临时目录
    import shutil as _sh
    _sh.rmtree(str(TEMP_DIR), ignore_errors=True)

    return ok({'saved': saved, 'skipped': skipped, 'errors': errors[:5]})


@app.route('/api/documents/batch-cancel', methods=['POST','OPTIONS'])
def docs_batch_cancel():
    """取消批量上传，清理临时文件"""
    if request.method == 'OPTIONS': return ok()
    import shutil
    batch_id = (request.json or {}).get('batch_id', '')
    if batch_id:
        TEMP_DIR = UPLOAD_DIR / 'tmp' / batch_id
        shutil.rmtree(str(TEMP_DIR), ignore_errors=True)
    return ok()


@app.route('/api/records/search', methods=['GET','OPTIONS'])
def records_search():
    """通用知产记录搜索（用于批量上传时手动匹配）"""
    if request.method == 'OPTIONS': return ok()
    q = request.args.get('q','').strip()
    if not q: return ok([])
    lk = f'%{q}%'
    with get_db() as db:
        patents = [dict(r) for r in db.execute(
            "SELECT id,'patent' AS item_type,title AS name,app_no FROM patents WHERE title LIKE ? OR app_no LIKE ? OR inventors LIKE ? LIMIT 10",
            [lk,lk,lk]).fetchall()]
        trademarks = [dict(r) for r in db.execute(
            "SELECT id,'trademark' AS item_type,name,app_no FROM trademarks WHERE name LIKE ? OR app_no LIKE ? LIMIT 10",
            [lk,lk]).fetchall()]
        copyrights = [dict(r) for r in db.execute(
            "SELECT id,'copyright' AS item_type,name,reg_no AS app_no FROM copyrights WHERE name LIKE ? OR reg_no LIKE ? LIMIT 10",
            [lk,lk]).fetchall()]
    return ok(patents + trademarks + copyrights)


# ══════════════ 待关联文件夹 ══════════════

@app.route('/api/watch/scan', methods=['GET','OPTIONS'])
def watch_scan():
    """扫描待关联文件夹，返回文件列表及自动匹配结果"""
    if request.method == 'OPTIONS': return ok()
    WATCH_DIR.mkdir(exist_ok=True)
    # Collect all files (recursively, up to 1 level of subdir)
    candidates = []
    for item in sorted(WATCH_DIR.rglob('*')):
        if item.is_file() and item.suffix.lower() in ALLOWED_EXT:
            rel = item.relative_to(WATCH_DIR)
            candidates.append((item, str(rel)))

    with get_db() as db:
        results = []
        for fpath, rel in candidates:
            item_type, record = _match_file(str(fpath), db)
            stat = fpath.stat()
            results.append({
                'rel_path': rel,
                'filename': fpath.name,
                'size': stat.st_size,
                'mtime': datetime.fromtimestamp(stat.st_mtime).strftime('%Y-%m-%d %H:%M'),
                'item_type': item_type,
                'item_id': record['id'] if record else None,
                'item_name': (record.get('title') or record.get('name') or '') if record else '',
                'item_app_no': (record.get('app_no') or record.get('reg_no') or '') if record else '',
                'suggested_doc_type': _doc_type_from_name(fpath.name),
                'matched': record is not None,
            })
    return ok({'files': results, 'watch_dir': str(WATCH_DIR), 'total': len(results)})


@app.route('/api/watch/import', methods=['POST','OPTIONS'])
def watch_import():
    """将待关联文件夹中选中的文件导入并关联，成功后从待关联目录移除"""
    if request.method == 'OPTIONS': return ok()
    items = (request.json or {}).get('items', [])
    if not items: return err('未指定文件', 400)
    imported, failed = 0, []
    with get_db() as db:
        for it in items:
            try:
                src = WATCH_DIR / it['rel_path']
                if not src.exists():
                    failed.append(it['rel_path'] + ' (不存在)'); continue
                if not it.get('item_id') or not it.get('item_type'):
                    failed.append(it['rel_path'] + ' (未指定关联目标)'); continue
                ext   = src.suffix.lower()
                new_fn = str(uuid.uuid4()) + ext
                dest   = UPLOAD_DIR / new_fn
                shutil.copy2(src, dest)
                mime = mimetypes.guess_type(src.name)[0] or 'application/octet-stream'
                db.execute("""INSERT INTO documents
                              (item_type,item_id,filename,original_name,file_size,mime_type,doc_type,description)
                              VALUES (?,?,?,?,?,?,?,?)""",
                           (it['item_type'], int(it['item_id']), new_fn, src.name,
                            dest.stat().st_size, mime,
                            it.get('doc_type','其他'), it.get('description','')))
                src.unlink()      # remove from watch folder
                imported += 1
            except Exception as e:
                failed.append(it['rel_path'] + f' ({e})')
    return ok({'imported': imported, 'failed': failed})


@app.route('/api/watch/path', methods=['GET','OPTIONS'])
def watch_path():
    """返回待关联文件夹的绝对路径"""
    if request.method == 'OPTIONS': return ok()
    WATCH_DIR.mkdir(exist_ok=True)
    return ok({'path': str(WATCH_DIR)})


@app.route('/api/export/package', methods=['POST','OPTIONS'])
def export_package():
    """统一导出：Excel（+可选附件 ZIP）"""
    if request.method == 'OPTIONS': return ok()
    try:
        d = request.get_json(force=True, silent=True) or {}
        data_type       = d.get('type','patents')
        filters         = d.get('filters',{})
        columns         = d.get('columns',[])
        with_attach     = bool(d.get('with_attachments', False))
        attach_types    = d.get('attachment_types', [])  # empty = all types

        def flike(v): return f'%{v}%'

        with get_db() as db:
            # ── Build query ──────────────────────────────────────────
            if data_type == 'patents':
                sql = "SELECT * FROM patents WHERE 1=1"; params = []
                q = filters.get('q','').strip()
                if q: sql += " AND (title LIKE ? OR app_no LIKE ? OR pub_no LIKE ? OR grant_no LIKE ? OR inventors LIKE ? OR owner LIKE ? OR ipc LIKE ?)"; params += [flike(q)]*7
                if filters.get('status'):         sql += " AND status=?";          params.append(filters['status'])
                if filters.get('type'):           sql += " AND type=?";            params.append(filters['type'])
                if filters.get('country'):        sql += " AND country LIKE ?";    params.append(flike(filters['country']))
                if filters.get('owner'):          sql += " AND owner LIKE ?";      params.append(flike(filters['owner']))
                if filters.get('inventor'):       sql += " AND inventors LIKE ?";  params.append(flike(filters['inventor']))
                if filters.get('fee_entity'):     sql += " AND fee_entity=?";      params.append(filters['fee_entity'])
                if filters.get('agent'):          sql += " AND agent LIKE ?";      params.append(flike(filters['agent']))
                if filters.get('pub_no'):         sql += " AND pub_no LIKE ?";     params.append(flike(filters['pub_no']))
                if filters.get('grant_no'):       sql += " AND grant_no LIKE ?";   params.append(flike(filters['grant_no']))
                if filters.get('app_date_from'):  sql += " AND app_date>=?";       params.append(filters['app_date_from'])
                if filters.get('app_date_to'):    sql += " AND app_date<=?";       params.append(filters['app_date_to'])
                if filters.get('grant_date_from'):sql += " AND grant_date>=?";     params.append(filters['grant_date_from'])
                if filters.get('grant_date_to'):  sql += " AND grant_date<=?";     params.append(filters['grant_date_to'])
                if filters.get('next_fee_from'):  sql += " AND next_fee_date>=?";  params.append(filters['next_fee_from'])
                if filters.get('next_fee_to'):    sql += " AND next_fee_date<=?";  params.append(filters['next_fee_to'])
                if filters.get('year_from'):      sql += " AND current_year>=?";   params.append(int(filters['year_from']))
                if filters.get('year_to'):        sql += " AND current_year<=?";   params.append(int(filters['year_to']))
                today = date.today().isoformat(); fd = filters.get('fee_due','')
                if fd == 'overdue': sql += " AND next_fee_date IS NOT NULL AND next_fee_date<?"; params.append(today)
                elif fd in ('7','30','90'): sql += " AND next_fee_date IS NOT NULL AND next_fee_date>=? AND next_fee_date<=date(?,'+'||? ||' days')"; params += [today, today, fd]
                rows = [row_to_dict(r) for r in db.execute(sql + " ORDER BY app_date DESC", params).fetchall()]

            elif data_type == 'trademarks':
                sql = "SELECT * FROM trademarks WHERE 1=1"; params = []
                q = filters.get('q','').strip()
                if q: sql += " AND (name LIKE ? OR app_no LIKE ? OR reg_no LIKE ? OR owner LIKE ? OR classes LIKE ?)"; params += [flike(q)]*5
                if filters.get('status'):            sql += " AND status=?";             params.append(filters['status'])
                if filters.get('type'):              sql += " AND type=?";               params.append(filters['type'])
                if filters.get('country'):           sql += " AND country LIKE ?";       params.append(flike(filters['country']))
                if filters.get('owner'):             sql += " AND owner LIKE ?";         params.append(flike(filters['owner']))
                if filters.get('classes'):           sql += " AND classes LIKE ?";       params.append(flike(filters['classes']))
                if filters.get('app_date_from'):     sql += " AND app_date>=?";          params.append(filters['app_date_from'])
                if filters.get('app_date_to'):       sql += " AND app_date<=?";          params.append(filters['app_date_to'])
                if filters.get('reg_date_from'):     sql += " AND reg_date>=?";          params.append(filters['reg_date_from'])
                if filters.get('reg_date_to'):       sql += " AND reg_date<=?";          params.append(filters['reg_date_to'])
                if filters.get('renewal_date_from'): sql += " AND renewal_date>=?";      params.append(filters['renewal_date_from'])
                if filters.get('renewal_date_to'):   sql += " AND renewal_date<=?";      params.append(filters['renewal_date_to'])
                rows = [row_to_dict(r) for r in db.execute(sql + " ORDER BY app_date DESC", params).fetchall()]

            else:  # copyrights
                sql = "SELECT * FROM copyrights WHERE 1=1"; params = []
                q = filters.get('q','').strip()
                if q: sql += " AND (name LIKE ? OR reg_no LIKE ? OR owner LIKE ?)"; params += [flike(q)]*3
                if filters.get('type'):               sql += " AND type=?";               params.append(filters['type'])
                if filters.get('owner'):              sql += " AND owner LIKE ?";         params.append(flike(filters['owner']))
                if filters.get('reg_date_from'):      sql += " AND reg_date>=?";          params.append(filters['reg_date_from'])
                if filters.get('reg_date_to'):        sql += " AND reg_date<=?";          params.append(filters['reg_date_to'])
                if filters.get('completion_date_from'):sql += " AND completion_date>=?";  params.append(filters['completion_date_from'])
                if filters.get('completion_date_to'): sql += " AND completion_date<=?";   params.append(filters['completion_date_to'])
                rows = [row_to_dict(r) for r in db.execute(sql + " ORDER BY reg_date DESC", params).fetchall()]

            if not rows:
                return err('筛选结果为空，请放宽条件', 404)

            # ── Build Excel ─────────────────────────────────────────
            import openpyxl as xl
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
            wb = xl.Workbook(); ws = wb.active
            type_label = {'patents':'专利','trademarks':'商标','copyrights':'软著'}.get(data_type,'数据')
            ws.title = type_label
            all_flds = {k: l for k, l in ALL_FIELDS[data_type]}
            export_cols = columns if columns else list(all_flds.keys())
            headers = [all_flds.get(k, k) for k in export_cols]
            ws.append(headers)
            hfill = PatternFill(fill_type='solid', fgColor='1F4E79')
            for ci, h in enumerate(headers, 1):
                cell = ws.cell(1, ci); cell.value = h
                cell.font = Font(bold=True, color='FFFFFF', name='Arial', size=10)
                cell.fill = hfill
                cell.alignment = Alignment(horizontal='center', vertical='center')
                ws.column_dimensions[get_column_letter(ci)].width = max(12, min(35, len(h)*2+8))
            ws.row_dimensions[1].height = 26
            alt = PatternFill(fill_type='solid', fgColor='EBF3FB')
            for ri, row in enumerate(rows, 2):
                ws.append([row.get(k, '') for k in export_cols])
                for ci in range(1, len(export_cols)+1):
                    cell = ws.cell(ri, ci); cell.font = Font(name='Arial', size=10)
                    cell.alignment = Alignment(vertical='center')
                    if ri % 2 == 0: cell.fill = alt
            ws.freeze_panes = 'A2'
            excel_buf = BytesIO(); wb.save(excel_buf); excel_buf.seek(0)
            fname_base = f'{type_label}数据_{date.today().isoformat()}'

            # ── Attach documents if requested ────────────────────────
            if with_attach:
                item_type_key = {'patents':'patent','trademarks':'trademark','copyrights':'copyright'}[data_type]
                ids = [r['id'] for r in rows]
                if ids:
                    ph = ','.join('?'*len(ids))
                    dt_filter = ""
                    dt_params = list(ids)
                    if attach_types:
                        ph2 = ','.join('?'*len(attach_types))
                        dt_filter = f" AND doc_type IN ({ph2})"
                        dt_params += attach_types
                    docs = db.execute(
                        f"SELECT * FROM documents WHERE item_type=? AND item_id IN ({ph}){dt_filter} ORDER BY item_id, doc_type",
                        [item_type_key] + dt_params
                    ).fetchall()
                else:
                    docs = []

                if docs:
                    # Build id→name map for folder naming
                    id_to_name = {}
                    for r in rows:
                        n = r.get('title') or r.get('name') or ''
                        app_no = r.get('app_no') or r.get('reg_no') or ''
                        id_to_name[r['id']] = f"{app_no}_{n}"[:40].replace('/','-').replace('\\','-')

                    zip_buf = BytesIO()
                    with zipfile.ZipFile(zip_buf, 'w', zipfile.ZIP_DEFLATED) as zf:
                        # Add Excel
                        excel_buf.seek(0)
                        zf.writestr(f'{fname_base}.xlsx', excel_buf.read())
                        # Add attachments grouped by record
                        name_count = {}
                        for doc in docs:
                            doc = dict(doc)
                            src = UPLOAD_DIR / doc['filename']
                            if not src.exists(): continue
                            folder = id_to_name.get(doc['item_id'], str(doc['item_id']))
                            oname  = doc['original_name'] or doc['filename']
                            zpath  = f'attachments/{folder}/{oname}'
                            if zpath in name_count:
                                name_count[zpath] += 1
                                base, ext = os.path.splitext(oname)
                                zpath = f'attachments/{folder}/{base}_{name_count[zpath]}{ext}'
                            else:
                                name_count[zpath] = 0
                            zf.write(src, zpath)
                    zip_buf.seek(0)
                    return send_file(zip_buf, mimetype='application/zip',
                                     as_attachment=True, download_name=f'{fname_base}.zip')

            # No attachments – return Excel only
            excel_buf.seek(0)
            return send_file(excel_buf,
                             mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                             as_attachment=True, download_name=f'{fname_base}.xlsx')
    except Exception as ex:
        import traceback; traceback.print_exc()
        return err(f'导出失败: {ex}', 500)



@app.route('/api/documents/bulk-delete', methods=['POST','OPTIONS'])
def bulk_delete_documents():
    if request.method == 'OPTIONS': return ok()
    ids = [int(x) for x in (request.json or {}).get('ids', [])]
    if not ids: return err('no ids', 400)
    with get_db() as db:
        rows = db.execute("SELECT filename FROM documents WHERE id IN ({})".format(','.join('?'*len(ids))), ids).fetchall()
        db.execute("DELETE FROM documents WHERE id IN ({})".format(','.join('?'*len(ids))), ids)
    for row in rows:
        try: (UPLOAD_DIR/row[0]).unlink(missing_ok=True)
        except: pass
    return ok({'deleted': len(rows)})


@app.route('/api/diagnostic')
def diagnostic():
    """诊断端点：返回服务器基本信息（仅本地访问）"""
    if request.remote_addr not in ('127.0.0.1', '::1'):
        return err('仅允许本地访问', 403)
    import sys
    return ok({
        'base_dir': str(BASE_DIR),
        'static_dir': str(STATIC_DIR),
        'index_exists': (STATIC_DIR/'index.html').exists(),
        'db_exists': DB_PATH.exists(),
        'python': sys.version,
        'port': 5678
    })


if __name__=='__main__':
    print("="*50); print("  知产管家 · IP Keeper 正在启动..."); print("="*50)
    init_db(); print("🌐 请在浏览器打开: http://localhost:5678"); print("="*50)
    app.run(host='127.0.0.1',port=5678,debug=False)
